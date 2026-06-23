from django.views.generic import ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import models
from apps.office.models import Note
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from datetime import datetime

class NoteListView(LoginRequiredMixin, ListView):
    model = Note
    template_name = 'office/note_list.html'
    context_object_name = 'notes'
    ordering = ['-created_at']

    paginate_by = 10

    def get_queryset(self):
        queryset = Note.objects.filter(is_deleted=False).select_related('employee').order_by('-created_at')
        
        # Filtering
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                models.Q(title__icontains=search_query) | 
                models.Q(description__icontains=search_query) |
                models.Q(employee__employee_name__icontains=search_query)
            )
            
        status_filter = self.request.GET.get('status', '')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        employee_id = self.request.GET.get('employee', '')
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)

        company_id = self.request.GET.get('company', '')
        if company_id:
            queryset = queryset.filter(employee__company_id=company_id)
            
        # Date range filtering
        from_date = self.request.GET.get('from_date', '')
        if from_date:
            queryset = queryset.filter(date__gte=from_date)
            
        to_date = self.request.GET.get('to_date', '')
        if to_date:
            queryset = queryset.filter(date__lte=to_date)
            
        return queryset

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            queryset = self.get_queryset()
            return self.export_to_excel(queryset)
        return super().get(request, *args, **kwargs)

    def export_to_excel(self, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Office_Notes.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'OFFICE NOTES'

        # Define styles
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = 'OFFICE NOTES EXPORT'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['Sl.No', 'Employee Name', 'Title', 'Description', 'Status', 'Date', 'Created At', 'Completed At']
        ws.append(headers)

        # Apply styles to header row (Row 2)
        for col_num, cell in enumerate(ws[2], 1):
            cell.fill = yellow_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment_center
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20

        # Add data
        for index, note in enumerate(queryset, 1):
            # Format created_at and completed_at datetime
            created_at = note.created_at.strftime('%d-%m-%Y %I:%M %p') if note.created_at else 'N/A'
            completed_at = note.completed_at.strftime('%d-%m-%Y %I:%M %p') if note.status == 'completed' and note.completed_at else 'N/A'
            
            row = [
                index,
                note.employee.employee_name if note.employee else 'N/A',
                note.title,
                note.description,
                note.get_status_display(),
                note.date,
                created_at,
                completed_at
            ]
            ws.append(row)

            # Set styles for the newly appended row
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = alignment_left

        wb.save(response)
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass employees and companies for dropdown filter
        from apps.authapp.models import Employee, Company
        context['employees'] = Employee.objects.filter(is_active=True, is_deleted=False, is_superuser=False, employee_type='office')
        context['companies'] = Company.objects.filter(is_active=True)
        return context

class NoteDetailView(LoginRequiredMixin, DetailView):
    model = Note
    template_name = 'office/note_detail.html'
    context_object_name = 'note'

    def get_queryset(self):
        return Note.objects.filter(is_deleted=False).select_related('employee')
