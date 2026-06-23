import json

from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.http import HttpResponseRedirect, JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from apps.authapp.models import Company, Employee
from apps.home.models import AttendanceCheck, BreakTimer, CompanyAnnouncement, Leave, LeaveBalance
from apps.home.views import calculate_leave_balance
from apps.profileapp.models import Document, VisaDetails, Vehicle, VehicleAssignment, ReportIssue, VehicleImage, TemporaryVehicleImage
from django.views.generic import ListView, CreateView, DeleteView, UpdateView
from django.urls import reverse_lazy
from datetime import datetime, timedelta, date
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.core.paginator import Paginator
from django.views.decorators.cache import never_cache
from django.db.models import Count, Sum, Avg, Q
from django.db.models.functions import TruncMonth
from django.core.mail import send_mail
from django.conf import settings
from apps.task.models import Task, DeliveryTask, Mechanic, ServiceAdvantage, ServiceTaskDax, ServiceDaxTypes, PLU, PartNumber, DeliveryNote, MechanicPartItem
from .models import Notification
from .forms import NotificationForm

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

from .models import Notification



class CompanyListView(LoginRequiredMixin, ListView):
    model = Company
    template_name = 'company_list.html'
    context_object_name = 'companies'
    
    def get_queryset(self):
        queryset = Company.objects.filter(is_deleted=False).order_by('company_name')
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(company_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(website__icontains=search_query)
            )
        return queryset

class CompanyCreateView(LoginRequiredMixin, CreateView):
    model = Company
    template_name = 'company_create.html'
    fields = ['company_name', 'address', 'phone', 'email', 'website', 'logo']
    
    def post(self, request, *args, **kwargs):
        try:
            company_name = request.POST.get('company_name')
            
            # Basic validation
            if not company_name:
                messages.error(request, "Company Name is required.")
                return render(request, self.template_name)
            
            company = Company(
                company_name=company_name,
                address=request.POST.get('address'),
                phone=request.POST.get('phone'),
                email=request.POST.get('email'),
                website=request.POST.get('website'),
            )
            
            if 'logo' in request.FILES:
                company.logo = request.FILES['logo']
                
            company.save()
            messages.success(request, f"Company '{company_name}' created successfully.")
            return redirect('company-list') 
            
        except Exception as e:
            messages.error(request, f"Error creating company: {str(e)}")
            return render(request, self.template_name)

class CompanyDetailView(LoginRequiredMixin, View):
    def get(self, request, pk):
        try:
            company = get_object_or_404(Company, pk=pk, is_deleted=False)
            employees = company.employees.filter(is_deleted=False).order_by('-id')
            context = {
                'company': company,
                'employees': employees,
            }
            return render(request, 'company_details.html', context)
        except Exception as e:
            messages.error(request, f'Error loading company details: {str(e)}')
            return redirect('company-list')

class CompanyUpdateView(LoginRequiredMixin, UpdateView):
    model = Company
    template_name = 'company_create.html'
    fields = ['company_name', 'address', 'phone', 'email', 'website', 'logo']
    context_object_name = 'company'
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            company_name = request.POST.get('company_name')
            
            if not company_name:
                messages.error(request, "Company Name is required.")
                return render(request, self.template_name, {'company': self.object})
            
            # Update fields
            self.object.company_name = company_name
            self.object.address = request.POST.get('address')
            self.object.phone = request.POST.get('phone')
            self.object.email = request.POST.get('email')
            self.object.website = request.POST.get('website')
            
            if 'logo' in request.FILES:
                self.object.logo = request.FILES['logo']
                
            self.object.save()
            messages.success(request, f"Company '{company_name}' updated successfully.")
            return redirect('company-list')
            
        except Exception as e:
            messages.error(request, f"Error updating company: {str(e)}")
            return render(request, self.template_name, {'company': self.object})

class CompanyDeleteView(LoginRequiredMixin, View):
    def post(self, request, company_id):
        try:
            company = Company.objects.get(id=company_id, is_deleted=False)
            company_name = company.company_name
            
            # Soft delete
            company.is_deleted = True
            company.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Company {company_name} has been deleted successfully.'
            })
            
        except Company.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Company not found.'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error deleting company: {str(e)}'
            }, status=500)


class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20

    def get_queryset(self):
        if getattr(self.request.user, 'role', '') in ['admin', 'super_admin']:
            queryset = Notification.objects.all()
        else:
            queryset = Notification.objects.filter(recipient=self.request.user)

        # Filtering
        if getattr(self.request.user, 'role', '') in ['admin', 'super_admin']:
            search_query = self.request.GET.get('search', '')
            if search_query:
                queryset = queryset.filter(
                    Q(title__icontains=search_query) |
                    Q(message__icontains=search_query) |
                    Q(recipient__employee_name__icontains=search_query)
                )

            notification_type = self.request.GET.get('type', '')
            if notification_type:
                queryset = queryset.filter(notification_type=notification_type)
            
            date_filter = self.request.GET.get('date', '')
            if date_filter:
                # Assuming created_at is DateTimeField
                queryset = queryset.filter(created_at__date=date_filter)

        return queryset.order_by('-created_at')
        

class SendNotificationView(LoginRequiredMixin, View):
    def get(self, request):
        if getattr(request.user, 'role', '') not in ['admin', 'super_admin']:
             messages.error(request, 'Permission denied.')
             return redirect('dashboard')
        
        form = NotificationForm()
        employees = Employee.objects.filter(is_active=True, is_deleted=False).order_by('employee_name')
        return render(request, 'send_notification.html', {'form': form, 'employees': employees})

    def post(self, request):
        if getattr(request.user, 'role', '') not in ['admin', 'super_admin']:
             messages.error(request, 'Permission denied.')
             return redirect('dashboard')

        form = NotificationForm(request.POST)
        employees = Employee.objects.filter(is_active=True, is_deleted=False).order_by('employee_name')
        if form.is_valid():
            title = form.cleaned_data['title']
            message = form.cleaned_data['message']
            notif_type = form.cleaned_data['notification_type']
            
            try:
                if notif_type == 'individual':
                    selected_employees = form.cleaned_data.get('employee')
                    if selected_employees:
                        # For multiple individual employees, we still create individual records
                        notifications = [
                            Notification(
                                recipient=emp,
                                title=title,
                                message=message,
                                notification_type=notif_type
                            ) for emp in selected_employees
                        ]
                        Notification.objects.bulk_create(notifications)
                        messages.success(request, f'Notification sent to {len(selected_employees)} employees successfully!')
                    else:
                        messages.warning(request, 'Please select at least one employee.')
                        return render(request, 'send_notification.html', {'form': form, 'employees': employees})

                elif notif_type == 'company':
                    company = form.cleaned_data.get('company')
                    if company:
                        Notification.objects.create(
                            target_company=company,
                            title=title,
                            message=message,
                            notification_type=notif_type
                        )
                        messages.success(request, f'Notification sent to {company.company_name} successfully!')
                    else:
                        messages.warning(request, 'Please select a company.')
                        return render(request, 'send_notification.html', {'form': form, 'employees': employees})

                elif notif_type == 'department':
                    department = form.cleaned_data.get('department')
                    if department:
                        Notification.objects.create(
                            target_department=department,
                            title=title,
                            message=message,
                            notification_type=notif_type
                        )
                        label = dict(Employee.EMPLOYEE_TYPE_CHOICES).get(department, department)
                        messages.success(request, f'Notification sent to {label} department successfully!')
                    else:
                        messages.warning(request, 'Please select a department.')
                        return render(request, 'send_notification.html', {'form': form, 'employees': employees})

                elif notif_type == 'common':
                    Notification.objects.create(
                        target_all=True,
                        title=title,
                        message=message,
                        notification_type=notif_type
                    )
                    messages.success(request, 'Common notification sent to all employees successfully!')
            
            except Exception as e:
                 messages.error(request, f'Error sending notification: {str(e)}')
                 return render(request, 'send_notification.html', {'form': form, 'employees': employees})

            return redirect('dashboard')
        
        return render(request, 'send_notification.html', {'form': form, 'employees': employees})



class AdminLogin(View):
    @method_decorator(never_cache)
    def get(self, request):
        if request.user.is_authenticated:
            return redirect('dashboard')
        return render(request, 'login.html')
    
    @method_decorator(never_cache)
    def post(self, request):
        # Get form data
        login_input = request.POST.get('EmployeeId') 
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')

        user = authenticate(request, username=login_input, password=password)
        
        if user is None:
            try:
                user_obj = Employee.objects.get(employeeId=login_input)
                user = authenticate(request, username=user_obj.employeeId, password=password)
            except Employee.DoesNotExist:
                user = None
        
        if user is not None:
            if self.can_user_login(user):
                login(request, user)
                
                if not remember_me:
                    request.session.set_expiry(0)

                if hasattr(user, 'get_full_name'):
                    request.session['user_full_name'] = user.get_full_name()
                elif hasattr(user, 'employee_name'):
                    request.session['user_full_name'] = user.employee_name or user.employeeId
                else:
                    request.session['user_full_name'] = user.employeeId
                
                if hasattr(user, 'employee'):
                    request.session['profile_image'] = user.employee.profile_image.url if user.employee.profile_image else None
                
                messages.success(request, 'Login successful!')
                return redirect('dashboard')
            else:
                messages.error(request, 'Access denied. Only admin users are allowed to login.')
                return render(request, 'login.html')
        else:
            messages.error(request, 'Invalid credentials. Please try again.')
            return render(request, 'login.html')
    
    def can_user_login(self, user):
        if hasattr(user, 'is_superuser') and user.is_superuser:
            return True
            
        if hasattr(user, 'is_deleted') and user.is_deleted:
            return False
        
        if hasattr(user, 'role'):
            return user.role in ['admin', 'super_admin']
        
        return False
    

class AdminLogout(View):
    @method_decorator(never_cache)
    def get(self, request):
        logout(request)
        return redirect('admin-login')

    
class EmployeeManage(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request):
        queryset = Employee.objects.filter(
            role__in=['super_admin', 'admin', 'employee'],  
            is_superuser=False,
            is_deleted=False
        ).order_by('-id')

        # Restrict to user's company if admin
        if getattr(request.user, 'role', '') == 'admin' and not request.user.is_superuser:
            if request.user.company:
                queryset = queryset.filter(company=request.user.company)


        # Search Filter
        search_query = request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(employeeId__icontains=search_query) |
                Q(employee_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(mobile_number__icontains=search_query)
            )

        department = request.GET.get('department')
        if department:
            queryset = queryset.filter(employee_type=department)

        company_id = request.GET.get('company')
        if company_id:
            queryset = queryset.filter(company_id=company_id)

        status = request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)

        # Pagination
        paginator = Paginator(queryset, 10)  
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Get companies for filter
        if getattr(request.user, 'role', '') == 'admin' and not request.user.is_superuser:
            if request.user.company:
                companies = Company.objects.filter(id=request.user.company.id, is_active=True, is_deleted=False)
            else:
                companies = Company.objects.none()
        else:
            companies = Company.objects.filter(is_active=True, is_deleted=False)

        context = {
            'employees': page_obj.object_list, 
            'page_obj': page_obj,            
            'is_paginated': page_obj.has_other_pages(),
            'department_choices': Employee.EMPLOYEE_TYPE_CHOICES,
            'companies': companies,
        }
        return render(request, 'customer.html', context)
    

class EmployeeCreate(LoginRequiredMixin, View):
    login_url = '/admin-login/'

    def get(self, request):
        context = {
            'role_choices': Employee.ROLE_CHOICES,
            'employee_type_choices': Employee.EMPLOYEE_TYPE_CHOICES,
            'document_choices': VisaDetails.DOCUMENT_TYPES,
            'companies': Company.objects.filter(is_active=True),
        }
        return render(request, 'create_customer.html', context)

    def post(self, request):
        try: 
            employee_id = request.POST.get('employeeId', '').strip()
            
            if not employee_id:
                messages.error(request, 'Employee ID is required!')
                return self.get(request)
            
            if Employee.objects.filter(employeeId=employee_id).exists():
                messages.error(request, f'Employee with ID {employee_id} already exists.')
                return self.get(request)
            
            # Get company instance
            company_id = request.POST.get('company')
            company_instance = None
            if company_id:
                try:
                    company_instance = Company.objects.get(id=company_id)
                except Company.DoesNotExist:
                    pass

            employee = Employee(
                employeeId=employee_id,
                employee_name=request.POST.get('employee_name', ''),
                role=request.POST.get('role', 'employee'),
                employee_type=request.POST.get('employee_type', 'service'),
                company=company_instance,
                mobile_number=request.POST.get('mobile_number', ''),
                email=request.POST.get('email', ''),
                home_address=request.POST.get('home_address', ''),
                nationality=request.POST.get('nationality', ''),
                emergency_contact_name=request.POST.get('emergency_contact_name', ''),
                emergency_contact_number=request.POST.get('emergency_contact_number', ''),
                emergency_contact_relation=request.POST.get('emergency_contact_relation', ''),
                is_active=request.POST.get('is_active') == 'true'
            )
            
            date_of_joining = request.POST.get('date_of_joining')
            
            if date_of_joining:
                 employee.date_joined = date_of_joining

            date_of_birth = request.POST.get('date_of_birth')
            
            if date_of_birth:
                employee.date_of_birth = date_of_birth
            
            if 'profile_pic' in request.FILES:
                employee.profile_pic = request.FILES['profile_pic']
            
            password = request.POST.get('password', '').strip()
            if not password:
                messages.error(request, 'Password is required!')
                return self.get(request)
            
            employee.set_password(password)
            
            employee.save()
            
            # Visa Details
            visa_details = VisaDetails(
                employee=employee,
                visa_expiry_date=request.POST.get('visa_expiry_date') or None,
                passport_number=request.POST.get('passport_number', ''),
                passport_expiry_date=request.POST.get('passport_expiry_date') or None,
                emirates_id_number=request.POST.get('emirates_id_number', ''),
                emirates_id_expiry=request.POST.get('emirates_id_expiry') or None,
            )
            visa_details.save()
            
            # Handle document uploads
            document_types = request.POST.getlist('document_types[]')
            
            for i, doc_type in enumerate(document_types):
                doc_files = request.FILES.getlist(f'document_files_{i}')
                for doc_file in doc_files:
                    if doc_type and doc_file:
                        try:
                            Document.objects.create(
                                visa_details=visa_details,
                                document_type=doc_type,
                                document_file=doc_file
                            )
                        except Exception as doc_error:
                            print(f"Error saving document {i+1}: {str(doc_error)}")  
            
            messages.success(request, f'Employee {employee.employeeId} created successfully!')
            return redirect('employee-manage')
            
        except Exception as e:
            print(f"Error in EmployeeCreate: {str(e)}")
            messages.error(request, f"Error creating employee: {str(e)}")
            context = {
                'role_choices': Employee.ROLE_CHOICES,
                'employee_type_choices': Employee.EMPLOYEE_TYPE_CHOICES,
                'document_choices': VisaDetails.DOCUMENT_TYPES,
                'companies': Company.objects.filter(is_active=True),
            }
            return render(request, 'create_customer.html', context)


class EmployeeDetailView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request, employee_id):
        try:
            employee = Employee.objects.select_related(
                'visa_details',
                'vehicle_assignment',
                'vehicle_assignment__vehicle',
            ).prefetch_related(
                'visa_details__documents',
                'vehicle_assignment__temporary_images',   
                'vehicle_assignment__vehicle__images'     
            ).get(
                id=employee_id,
                is_superuser=False,
                is_deleted=False
            )

            # Fetch and recalculate leave balances to ensure fresh data
            from datetime import datetime
            current_year = timezone.now().year
            current_month = timezone.now().month
            
            # Recalculate current month's balance
            calculate_leave_balance(employee, current_year, current_month)
            
            # Also recalculate for any months with approved leaves to ensure accuracy
            employee_leaves = Leave.objects.filter(
                employee=employee,
                status='approved'
            ).distinct()
            
            for leave in employee_leaves:
                try:
                    leave_start_date = datetime.strptime(leave.start_date, '%Y-%m-%d')
                    calculate_leave_balance(employee, leave_start_date.year, leave_start_date.month)
                except (ValueError, TypeError):
                    pass
            
            leave_balances = LeaveBalance.objects.filter(employee=employee).order_by('-year', '-month')
            current_balance = leave_balances.first()
            
            # Calculate year-based totals for display
            from decimal import Decimal
            annual_entitlement = Decimal('30')  # 2.5 × 12 months
            
            # Calculate total used this year (sum of all approved leaves)
            total_used_this_year = Decimal('0')
            year_start = datetime(current_year, 1, 1)
            year_end = datetime(current_year, 12, 31)
            
            approved_leaves_this_year = Leave.objects.filter(
                employee=employee,
                status='approved'
            )
            
            for leave in approved_leaves_this_year:
                try:
                    leave_start = datetime.strptime(leave.start_date, '%Y-%m-%d')
                    # Count if leave starts in current year
                    if leave_start.year == current_year:
                        total_used_this_year += (leave.total_days or Decimal('0'))
                except (ValueError, TypeError):
                    pass
            
            year_remaining = annual_entitlement - total_used_this_year
            
            # Calculate carried forward from previous year (max 30 days)
            carried_forward = Decimal('0')
            prev_year = current_year - 1
            prev_year_dec = LeaveBalance.objects.filter(
                employee=employee,
                year=prev_year,
                month=12
            ).first()
            
            if prev_year_dec and prev_year_dec.remaining > 0:
                carried_forward = min(prev_year_dec.remaining, Decimal('30'))
            
            # Total available = annual entitlement + carried forward
            total_available_with_carry = annual_entitlement + carried_forward
            year_remaining_with_carry = total_available_with_carry - total_used_this_year

            context = {
                'employee': employee,
                'vehicles': Vehicle.objects.all(),
                'fuel_types': Vehicle.FUEL_TYPES,
                'leave_balances': leave_balances,
                'current_leave_balance': current_balance,
                # Year-based totals for display
                'annual_entitlement': annual_entitlement,
                'carried_forward': carried_forward,
                'total_available': total_available_with_carry,
                'year_used_total': total_used_this_year,
                'year_remaining': year_remaining_with_carry,
                'current_year': current_year,
            }
            return render(request, 'employee_details.html', context)

        except Employee.DoesNotExist:
            messages.error(request, 'Employee not found')
            return redirect('employee-manage')
            

class EmployeeEditView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request, employee_id):
        try:
            employee = Employee.objects.select_related('visa_details').prefetch_related('visa_details__documents').get(
                id=employee_id,
                is_superuser=False,
                is_deleted=False
            )

            visa_details = None
            if hasattr(employee, 'visa_details'):
                visa_details = employee.visa_details

            context = {
                'employee': employee,
                'visa_details': visa_details,
                'role_choices': Employee.ROLE_CHOICES,
                'employee_type_choices': Employee.EMPLOYEE_TYPE_CHOICES,
                'document_choices': VisaDetails.DOCUMENT_TYPES,
                'companies': Company.objects.filter(is_active=True),
            }

            return render(request, 'edit_customer.html', context)
            
        except Employee.DoesNotExist:
            messages.error(request, 'Employee not found.')
            return redirect('employee-manage')
        except Exception as e:
            messages.error(request, f'Error loading employee: {str(e)}')
            return redirect('employee-manage')
    
    def post(self, request, employee_id):
        try:
            employee = Employee.objects.get(id=employee_id, is_superuser=False, is_deleted=False)
            
            # Update basic employee information
            employee.employeeId = request.POST.get('employeeId', employee.employeeId)
            employee.employee_name = request.POST.get('employee_name', employee.employee_name)
            employee.role = request.POST.get('role', employee.role)
            employee.employee_type = request.POST.get('employee_type', employee.employee_type)
            
            company_id = request.POST.get('company')
            if company_id:
                try:
                    employee.company = Company.objects.get(id=company_id)
                except Company.DoesNotExist:
                    pass
            
            employee.mobile_number = request.POST.get('mobile_number', employee.mobile_number)
            employee.email = request.POST.get('email', employee.email)
            employee.home_address = request.POST.get('home_address', employee.home_address)
            employee.nationality = request.POST.get('nationality', employee.nationality)
            employee.emergency_contact_name = request.POST.get('emergency_contact_name', employee.emergency_contact_name)
            employee.emergency_contact_number = request.POST.get('emergency_contact_number', employee.emergency_contact_number)
            employee.emergency_contact_relation = request.POST.get('emergency_contact_relation', employee.emergency_contact_relation)
            employee.is_active = request.POST.get('is_active') == 'true'
            
            date_of_birth = request.POST.get('date_of_birth')
            if date_of_birth:
                employee.date_of_birth = date_of_birth
            
            date_of_joining = request.POST.get('date_of_joining')
            if date_of_joining:
                employee.date_joined = date_of_joining
            
            if 'profile_pic' in request.FILES:
                employee.profile_pic = request.FILES['profile_pic']
            
            employee.save()
            
            # Update Visa Details
            visa_details, created = VisaDetails.objects.get_or_create(employee=employee)
            visa_details.passport_number = request.POST.get('passport_number', visa_details.passport_number)
            visa_details.emirates_id_number = request.POST.get('emirates_id_number', visa_details.emirates_id_number)
            
            visa_expiry_date = request.POST.get('visa_expiry_date')
            passport_expiry_date = request.POST.get('passport_expiry_date')
            emirates_id_expiry = request.POST.get('emirates_id_expiry')
            
            if visa_expiry_date:
                visa_details.visa_expiry_date = visa_expiry_date
            if passport_expiry_date:
                visa_details.passport_expiry_date = passport_expiry_date
            if emirates_id_expiry:
                visa_details.emirates_id_expiry = emirates_id_expiry
            
            visa_details.save()
            
            # Handle document uploads
            document_types = request.POST.getlist('document_types[]')
            
            for i, doc_type in enumerate(document_types):
                doc_files = request.FILES.getlist(f'document_files_{i}')
                for doc_file in doc_files:
                    if doc_type and doc_file:
                        Document.objects.create(
                            visa_details=visa_details,
                            document_type=doc_type,
                            document_file=doc_file
                        )
            
            messages.success(request, f'Employee {employee.employeeId} updated successfully!')
            return redirect('employee-manage')
            
        except Employee.DoesNotExist:
            messages.error(request, 'Employee not found.')
            return redirect('employee-manage')
        except Exception as e:
            messages.error(request, f'Error updating employee: {str(e)}')
            return self.get(request, employee_id)
        

class EmployeeDeleteView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def delete(self, request, employee_id):
        try:
            employee = Employee.objects.get(id=employee_id, is_superuser=False, is_deleted=False)
            
            # Store employee info for success message
            employee_id_str = employee.employeeId
            employee_name = employee.employee_name
            
            # Delete the employee
            employee.is_deleted = True
            employee.save()
            
            # Return success response
            return JsonResponse({
                'success': True,
                'message': f'Employee {employee_name} ({employee_id_str}) has been deleted successfully.'
            })
            
        except Employee.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Employee not found.'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error deleting employee: {str(e)}'
            }, status=500)


class AssignVehicleView(LoginRequiredMixin, View):
    def post(self, request, employee_id):
        # Role Validation
        if getattr(request.user, 'role', '') not in ['admin', 'super_admin']:
            messages.error(request, 'You do not have permission to perform this action.')
            return redirect('employee-manage')

        try:
            employee = get_object_or_404(Employee, id=employee_id, is_deleted=False)
            vehicle_id = request.POST.get('vehicle')
            

            # Create or update assignment
            assignment, created = VehicleAssignment.objects.get_or_create(employee=employee)
            
            if vehicle_id:
                vehicle = get_object_or_404(Vehicle, id=vehicle_id)
                assignment.vehicle = vehicle
            
            assignment.current_vehicle_assigned_date = request.POST.get('current_vehicle_assigned_date') or None
            assignment.current_vehicle_assigned_time = request.POST.get('current_vehicle_assigned_time') or None
            assignment.current_vehicle_ending_date = request.POST.get('current_vehicle_ending_date') or None
            assignment.current_vehicle_ending_time = request.POST.get('current_vehicle_ending_time') or None
            
            assignment.save()
            
            messages.success(request, 'Vehicle assigned successfully.')
            return redirect('employee-details', employee_id=employee_id)
            
        except Exception as e:
            messages.error(request, f"Error assigning vehicle: {str(e)}")
            return redirect('employee-details', employee_id=employee_id)


class AddVehicleView(LoginRequiredMixin, View):
    def post(self, request):
        # Role Validation
        if getattr(request.user, 'role', '') not in ['admin', 'super_admin']:
            messages.error(request, 'You do not have permission to perform this action.')
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

        try:
            vehicle_number = request.POST.get('vehicle_number', '').strip()
            
            # Duplicate Validation
            if Vehicle.objects.filter(vehicle_number__iexact=vehicle_number).exists():
                messages.error(request, f"Vehicle with number {vehicle_number} already exists.")
                return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

            vehicle = Vehicle(
                vehicle_number=vehicle_number,
                model=request.POST.get('model'),
                vehicle_type=request.POST.get('vehicle_type'),
                fuel_type=request.POST.get('fuel_type'),
                insurance_expiry_date=request.POST.get('insurance_expiry_date'),
            )
            
            vehicle.save()

            # Handle multiple images
            images = request.FILES.getlist('vehicle_images')
            print(images,"this are the imagesssss")
            for image in images:
                VehicleImage.objects.create(vehicle=vehicle, image=image)
            
            messages.success(request, 'Vehicle added successfully.')
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
            
        except Exception as e:
            messages.error(request, f"Error adding vehicle: {str(e)}")
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


class VehicleListView(LoginRequiredMixin, View):
    def get(self, request):
        vehicles = Vehicle.objects.all().order_by('-created_at')
        context = {
            'vehicles': vehicles,
            'fuel_types': Vehicle.FUEL_TYPES,
        }
        return render(request, 'vehicle_list.html', context)


class VehicleUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if getattr(request.user, 'role', '') not in ['admin', 'super_admin']:
            messages.error(request, 'You do not have permission to perform this action.')
            return redirect('vehicle-list')

        try:
            vehicle = get_object_or_404(Vehicle, pk=pk)
            vehicle_number = request.POST.get('vehicle_number', '').strip()
            
            # Duplicate check (excluding current vehicle)
            if vehicle_number and vehicle_number.lower() != vehicle.vehicle_number.lower():
                if Vehicle.objects.filter(vehicle_number__iexact=vehicle_number).exclude(pk=pk).exists():
                    messages.error(request, f"Vehicle with number {vehicle_number} already exists.")
                    return redirect('vehicle-list')
                vehicle.vehicle_number = vehicle_number

            vehicle.model = request.POST.get('model')
            vehicle.vehicle_type = request.POST.get('vehicle_type')
            vehicle.fuel_type = request.POST.get('fuel_type')
            vehicle.insurance_expiry_date = request.POST.get('insurance_expiry_date')
            
            vehicle.save()

            # Handle multiple images
            images = request.FILES.getlist('vehicle_images')
            for image in images:
                VehicleImage.objects.create(vehicle=vehicle, image=image)

            messages.success(request, 'Vehicle updated successfully.')
            
        except Exception as e:
            messages.error(request, f"Error updating vehicle: {str(e)}")
            
        return redirect('vehicle-list')


class VehicleDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if getattr(request.user, 'role', '') not in ['admin', 'super_admin']:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
            messages.error(request, 'You do not have permission to perform this action.')
            return redirect('vehicle-list')

        try:
            vehicle = Vehicle.objects.get(pk=pk)
            vehicle.delete()
            # Return JSON for AJAX calls if needed, or redirect
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True})
            
            messages.success(request, 'Vehicle deleted successfully.')
            
        except Vehicle.DoesNotExist:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Vehicle not found.'}, status=404)
            messages.error(request, 'Vehicle not found.')
            
        except Exception as e:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                 return JsonResponse({'success': False, 'error': str(e)}, status=500)
            messages.error(request, f"Error deleting vehicle: {str(e)}")
            
        return redirect('vehicle-list')


class DeleteVehicleImageView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if getattr(request.user, 'role', '') not in ['admin', 'super_admin']:
            return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

        try:
            image = get_object_or_404(VehicleImage, pk=pk)
            image.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    
class DailyAttendanceView(LoginRequiredMixin,View):
    def get(self, request):
        date_filter = request.GET.get('date')
        if date_filter:
            target_date_str = date_filter
            try:
                target_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            except ValueError:
                target_date = timezone.localdate()
                target_date_str = target_date.isoformat()
        else:
            target_date = timezone.localdate()
            target_date_str = target_date.isoformat()
        employees = Employee.objects.filter(
            is_active=True, 
            is_deleted=False,
            is_superuser=False,
            is_staff=False
        )
        
        # Restrict to user's company if admin
        if getattr(request.user, 'role', '') == 'admin' and not request.user.is_superuser:
            if request.user.company:
                employees = employees.filter(company=request.user.company)
        
        # Filter by Company
        company_id = request.GET.get('company')
        if company_id:
            employees = employees.filter(company_id=company_id)
            
        # Filter by Department (Employee Type)
        department = request.GET.get('department')
        if department:
            employees = employees.filter(employee_type=department)
            
        employees = employees.exclude(role__in=['super_admin'])
                
        daily_data = []
        present_count = 0
        absent_count = 0
        
        # Get context data for filters
        if getattr(request.user, 'role', '') == 'admin' and not request.user.is_superuser:
            if request.user.company:
                companies = Company.objects.filter(id=request.user.company.id, is_active=True)
            else:
                companies = Company.objects.none()
        else:
            companies = Company.objects.filter(is_active=True)
        departments = Employee.EMPLOYEE_TYPE_CHOICES
        
        for employee in employees:
            # Use string date
            attendance_records = AttendanceCheck.objects.filter(
                employee=employee,
                check_date=target_date_str
            ).order_by('check_time')
            
            break_records = BreakTimer.objects.filter(
                employee=employee,
                date=target_date_str
            )
            
            # Determine presence based on having at least one check-in
            is_present = attendance_records.filter(check_type='in').exists()
            if is_present:
                present_count += 1
            else:
                absent_count += 1
            
            total_break_minutes = 0
            total_extend_count = 0
            for break_record in break_records:
                total_extend_count += break_record.extend_count
                if break_record.break_end_time:
                    try:
                        start_time = datetime.strptime(break_record.break_start_time, '%H:%M:%S')
                    except ValueError:
                        try:
                            start_time = datetime.strptime(break_record.break_start_time, '%H:%M')
                        except ValueError:
                            continue

                    try:
                        end_time = datetime.strptime(break_record.break_end_time, '%H:%M:%S')
                    except ValueError:
                        try:
                            end_time = datetime.strptime(break_record.break_end_time, '%H:%M')
                        except ValueError:
                            continue

                    break_duration = (end_time - start_time).total_seconds() / 60
                    total_break_minutes += break_duration
            
            working_hours = "N/A"
            # Calculate Total Working Hours from Multiple Sessions
            total_working_minutes = 0
            sessions = []
            
            # Filter valid records for calculation
            # We need to process them in chronological order
            attendance_list = list(attendance_records)
            
            i = 0
            while i < len(attendance_list):
                current_record = attendance_list[i]
                
                if current_record.check_type == 'in':
                    start_time_obj = None
                    try:
                        start_time_obj = datetime.strptime(current_record.check_time, '%H:%M:%S')
                    except ValueError:
                        try:
                            start_time_obj = datetime.strptime(current_record.check_time, '%H:%M')
                        except ValueError:
                            pass
                            
                    if start_time_obj:
                        session_data = {
                            'check_in': current_record.check_time,
                            'check_out': None,
                        }
                        
                        # Find corresponding check-out
                        end_time_obj = None
                        if i + 1 < len(attendance_list) and attendance_list[i+1].check_type == 'out':
                            next_record = attendance_list[i+1]
                            session_data['check_out'] = next_record.check_time
                            try:
                                end_time_obj = datetime.strptime(next_record.check_time, '%H:%M:%S')
                            except ValueError:
                                try:
                                    end_time_obj = datetime.strptime(next_record.check_time, '%H:%M')
                                except ValueError:
                                    pass
                            i += 2 # Move past this pair
                        else:
                            # No checkout found (Open session)
                            # Handle auto-checkout logic for display purposes
                            # If date is today and time is past 10 PM (22:00), or date is in past
                            is_today = (target_date == timezone.localdate())
                            current_now = timezone.now()
                            
                            cutoff_time = datetime.strptime("22:00:00", "%H:%M:%S").time()
                            
                            should_auto_close = False
                            if target_date < timezone.localdate():
                                should_auto_close = True
                            elif is_today and current_now.time() > cutoff_time:
                                should_auto_close = True
                                
                            if should_auto_close:
                                # Use same date components as start_time but set to 22:00
                                end_time_obj = start_time_obj.replace(hour=22, minute=0, second=0)
                                if start_time_obj > end_time_obj:
                                     end_time_obj = start_time_obj # 0 duration
                            
                            i += 1 # Move to next record (if any)
                        
                        sessions.append(session_data)

                        if end_time_obj:
                            # Handle crossing midnight
                            if end_time_obj < start_time_obj:
                                end_time_obj += timedelta(days=1)
                                
                            duration = (end_time_obj - start_time_obj).total_seconds() / 60
                            total_working_minutes += duration

                else:
                    # Starting with 'out' record (orphan), skip
                    i += 1

            # Subtract total break time
            actual_working_minutes = total_working_minutes - total_break_minutes
            if actual_working_minutes < 0: actual_working_minutes = 0
            
            working_hours = f"{int(actual_working_minutes // 60)}h {int(actual_working_minutes % 60)}m" if total_working_minutes > 0 else "N/A"
            
            # Get locations
            check_in_record = attendance_records.filter(check_type='in').first()
            check_in_location = check_in_record.location if check_in_record and check_in_record.location else 'N/A'
            check_out_record = attendance_records.filter(check_type='out').last()
            check_out_location = check_out_record.location if check_out_record and check_out_record.location else 'N/A'
            
            daily_data.append({
                'employee': employee,
                'sessions': sessions,
                'break_count': break_records.count(),
                'extend_count': total_extend_count,
                'total_break_time': f"{int(total_break_minutes // 60)}h {int(total_break_minutes % 60)}m",
                'working_hours': working_hours,
                'status': 'Present' if is_present else 'Absent',
                'check_in_location': check_in_location,
                'check_out_location': check_out_location
            })
        
        if request.GET.get('export') == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from django.http import HttpResponse
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Daily Attendance"
            
            # Define styles
            header_font = Font(bold=True, color="000000")
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Write headers with yellow background
            headers = ['SI NO', 'Date', 'Employee Name', 'Employee ID', 'Status', 'Check In', 'Check Out', 'Breaks', 'Extend Count', 'Total Break Time', 'Working Hours', 'Check In Location', 'Check Out Location']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Write data rows
            for row_num, data in enumerate(daily_data, 2):
                # Format sessions
                check_ins = " | ".join([str(s['check_in']) for s in data['sessions'] if s['check_in']])
                check_outs = " | ".join([str(s['check_out']) for s in data['sessions'] if s['check_out']]) if data['sessions'] else "-"
                
                ws.cell(row=row_num, column=1, value=row_num - 1)
                ws.cell(row=row_num, column=2, value=target_date)
                ws.cell(row=row_num, column=3, value=data['employee'].employee_name or "")
                ws.cell(row=row_num, column=4, value=data['employee'].employeeId)
                ws.cell(row=row_num, column=5, value=data['status'])
                ws.cell(row=row_num, column=6, value=check_ins)
                ws.cell(row=row_num, column=7, value=check_outs)
                ws.cell(row=row_num, column=8, value=data['break_count'])
                ws.cell(row=row_num, column=9, value=data['extend_count'])
                ws.cell(row=row_num, column=10, value=data['total_break_time'])
                ws.cell(row=row_num, column=11, value=data['working_hours'])
                ws.cell(row=row_num, column=12, value=data['check_in_location'])
                ws.cell(row=row_num, column=13, value=data['check_out_location'])
                
                # Add borders to data cells
                for col_num in range(1, 14):
                    ws.cell(row=row_num, column=col_num).border = border
            
            # Adjust column widths
            column_widths = [8, 12, 20, 15, 10, 20, 20, 8, 12, 15, 12, 20, 20]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
            
            # Save to response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="daily_attendance_{target_date}.xlsx"'
            wb.save(response)
            
            return response

        context = {
            'selected_date': target_date,
            'daily_data': daily_data,
            'total_employees': len(daily_data),
            'present_count': present_count,
            'absent_count': absent_count,
            'companies': companies,
            'departments': departments,
            'selected_company': int(company_id) if company_id else None,
            'selected_department': department,
        }
        
        return render(request, 'daily_attendance.html', context)
    

class AttendanceListView(LoginRequiredMixin,View):
    def get(self, request):
        # Date Range Filter
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Default to today if no date range provided
        if not start_date and not end_date:
            today_str = timezone.localdate().isoformat()
            start_date = today_str
            end_date = today_str
        
        # Generate list of dates for the range
        date_list = []
        if start_date and end_date:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            current_date = start
            while current_date <= end:
                date_list.append(current_date.isoformat())
                current_date += timedelta(days=1)
        
        # Get all active employees based on filters
        employees = Employee.objects.filter(
            is_active=True, 
            is_superuser=False,
            is_staff=False,
            is_deleted=False
        )
        
        # Restrict to user's company if admin
        if getattr(request.user, 'role', '') == 'admin' and not request.user.is_superuser:
            if request.user.company:
                employees = employees.filter(company=request.user.company)
        
        # Apply company filter
        company_filter = request.GET.get('company')
        if company_filter:
            employees = employees.filter(company_id=company_filter)
        
        # Apply employee filter
        employee_filter = request.GET.get('employee')
        if employee_filter:
            employees = employees.filter(
                Q(employeeId__icontains=employee_filter) |
                Q(employee_name__icontains=employee_filter)
            )
        
        # Get all attendance records for the date range and employees
        attendance_records = AttendanceCheck.objects.filter(
            employee__in=employees,
            check_date__gte=start_date,
            check_date__lte=end_date
        ).select_related('employee', 'employee__company')
        
        # Apply check type filter
        check_type = request.GET.get('check_type')
        if check_type in ['in', 'out']:
            attendance_records = attendance_records.filter(check_type=check_type)
        
        # Sort attendance records
        sort_by = request.GET.get('sort_by', '-date')
        ordering = []
        if sort_by == 'date_asc':
            ordering = ['check_date', 'employee__employeeId', 'check_time']
        elif sort_by == 'date_desc':
            ordering = ['-check_date', 'employee__employeeId', 'check_time']
        elif sort_by == 'name_asc':
            ordering = ['employee__employee_name', '-check_date', 'check_time']
        elif sort_by == 'name_desc':
            ordering = ['-employee__employee_name', '-check_date', 'check_time']
        else:
            ordering = ['-check_date', 'employee__employeeId', 'check_time']
        
        attendance_records = attendance_records.order_by(*ordering)
        
        # Create attendance lookup dictionary
        attendance_lookup = {}
        for record in attendance_records:
            key = f"{record.employee.id}_{record.check_date}"
            if key not in attendance_lookup:
                attendance_lookup[key] = []
            attendance_lookup[key].append(record)
        
        # Create comprehensive attendance data including absent employees
        grouped_records = []
        si_no = 1
        
        for employee in employees:
            for date_str in date_list:
                try:
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    date_obj = date_str
                
                key = f"{employee.id}_{date_str}"
                records = attendance_lookup.get(key, [])
                
                # Calculate working hours
                working_hours = '0h 0m'
                if records:
                    check_in_records = [r for r in records if r.check_type == 'in']
                    check_out_records = [r for r in records if r.check_type == 'out']
                    
                    if check_in_records and check_out_records:
                        # Get first check-in and last check-out
                        first_check_in = min(check_in_records, key=lambda x: x.check_time or '00:00:00')
                        last_check_out = max(check_out_records, key=lambda x: x.check_time or '00:00:00')
                        
                        if first_check_in.check_time and last_check_out.check_time:
                            try:
                                # Try %H:%M:%S format first, then %H:%M
                                try:
                                    time_in = datetime.strptime(first_check_in.check_time, '%H:%M:%S')
                                except ValueError:
                                    time_in = datetime.strptime(first_check_in.check_time, '%H:%M')
                                
                                try:
                                    time_out = datetime.strptime(last_check_out.check_time, '%H:%M:%S')
                                except ValueError:
                                    time_out = datetime.strptime(last_check_out.check_time, '%H:%M')
                                
                                # Calculate difference
                                diff = time_out - time_in
                                total_minutes = int(diff.total_seconds() // 60)
                                hours = total_minutes // 60
                                minutes = total_minutes % 60
                                working_hours = f"{hours}h {minutes}m"
                            except (ValueError, TypeError):
                                working_hours = '0h 0m'
                
                grouped_records.append({
                    'si_no': si_no,
                    'employee': employee,
                    'date': date_obj,
                    'date_str': date_str,
                    'records': records,
                    'status': 'Present' if records else 'Absent',
                    'working_hours': working_hours
                })
                si_no += 1
        
        # Apply status filter
        status_filter = request.GET.get('status')
        if status_filter == 'present':
            grouped_records = [gr for gr in grouped_records if gr['status'] == 'Present']
        elif status_filter == 'absent':
            grouped_records = [gr for gr in grouped_records if gr['status'] == 'Absent']
        
        # Sort grouped records
        if sort_by == 'name_asc':
            grouped_records.sort(key=lambda x: (x['employee'].employee_name or '', x['date_str']))
        elif sort_by == 'name_desc':
            grouped_records.sort(key=lambda x: (x['employee'].employee_name or '', x['date_str']), reverse=True)
        elif sort_by == 'date_asc':
            grouped_records.sort(key=lambda x: (x['date_str'], x['employee'].employeeId or ''))
        elif sort_by == 'date_desc':
            grouped_records.sort(key=lambda x: (x['date_str'], x['employee'].employeeId or ''), reverse=True)
        else:  # -date (default)
            grouped_records.sort(key=lambda x: (x['date_str'], x['employee'].employeeId or ''), reverse=True)
        
        # Recalculate serial numbers after filtering and sorting
        for idx, record in enumerate(grouped_records, start=1):
            record['si_no'] = idx
        
        # Pagination
        paginator = Paginator(grouped_records, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Get filter values for context
        current_employee = request.GET.get('employee', '')
        current_company = company_filter
        current_check_type = request.GET.get('check_type', '')
        current_sort = request.GET.get('sort_by', '')
        current_status = request.GET.get('status', '')
        
        # Get companies for dropdown
        if getattr(request.user, 'role', '') == 'admin' and not request.user.is_superuser:
            if request.user.company:
                companies = Company.objects.filter(id=request.user.company.id, is_active=True)
            else:
                companies = Company.objects.none()
        else:
            companies = Company.objects.filter(is_active=True)
        
        # Get all employees for dropdown (filter by company if selected)
        dropdown_employees = Employee.objects.filter(
            is_active=True, 
            is_superuser=False,
            is_staff=False,
            is_deleted=False
        )
        if getattr(request.user, 'role', '') == 'admin' and not request.user.is_superuser:
            if request.user.company:
                dropdown_employees = dropdown_employees.filter(company=request.user.company)
        if company_filter:
            dropdown_employees = dropdown_employees.filter(company_id=company_filter)

        context = {
            'grouped_records': page_obj,
            'page_obj': page_obj,
            'is_paginated': page_obj.has_other_pages(),
            'start_date': start_date,
            'end_date': end_date,
            'current_employee': current_employee,
            'current_company': current_company,
            'current_check_type': current_check_type,
            'current_sort': current_sort,
            'current_status': current_status,
            'employees': dropdown_employees,
            'companies': companies,
        }
        
        # Check for Excel export request
        if request.GET.get('export') == 'excel':
            return self.export_attendance_to_excel(request, grouped_records)
        
        return render(request, 'attendance_list.html', context)
    

    def export_attendance_to_excel(self, request, grouped_records):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Records"

        # Styles with YELLOW headers
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # New headers as requested
        headers = ['SI No', 'Date', 'Employee ID', 'Employee Name', 'Check In Time', 'Check Out Time', 
                  'Check In Location', 'Check Out Location', 'Status', 'Total Break Time', 'Breaks', 'Extend Count', 'Working Hours']
        ws.append(headers)

        # Apply header styling with yellow background
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = border

        # Process grouped data and write to Excel
        for data in grouped_records:
            employee = data['employee']
            check_date = data['date_str']
            records = data['records']
            
            # Extract check-in and check-out times and locations
            check_in_time = None
            check_out_time = None
            check_in_location = "N/A"
            check_out_location = "N/A"
            
            for record in records:
                if record.check_type == 'in':
                    if not check_in_time:
                        check_in_time = record.check_time
                    if check_in_location == "N/A" and record.location:
                        check_in_location = record.location
                elif record.check_type == 'out':
                    if not check_out_time:
                        check_out_time = record.check_time
                    if check_out_location == "N/A" and record.location:
                        check_out_location = record.location
            
            # Get working hours and break history for the employee on this date
            from apps.home.models import BreakTimer
            
            # Ensure we have ALL records (both in and out) for working hours calculation
            # data['records'] might be filtered by check_type from the UI filter
            has_in = any(r.check_type == 'in' for r in records)
            has_out = any(r.check_type == 'out' for r in records)
            if records and (not has_in or not has_out):
                # Re-fetch unfiltered records for this employee+date
                calc_records = list(AttendanceCheck.objects.filter(
                    employee=employee,
                    check_date=check_date
                ).order_by('check_time'))
            else:
                calc_records = list(records)
            
            total_working_minutes = 0
            all_records = sorted(calc_records, key=lambda r: r.check_time or '00:00:00')
            
            i = 0
            while i < len(all_records):
                current_record = all_records[i]
                
                if current_record.check_type == 'in':
                    start_time_obj = None
                    try:
                        start_time_obj = datetime.strptime(current_record.check_time, '%H:%M:%S')
                    except (ValueError, TypeError):
                        try:
                            start_time_obj = datetime.strptime(current_record.check_time, '%H:%M')
                        except (ValueError, TypeError):
                            pass
                    
                    if start_time_obj:
                        end_time_obj = None
                        if i + 1 < len(all_records) and all_records[i + 1].check_type == 'out':
                            next_record = all_records[i + 1]
                            try:
                                end_time_obj = datetime.strptime(next_record.check_time, '%H:%M:%S')
                            except (ValueError, TypeError):
                                try:
                                    end_time_obj = datetime.strptime(next_record.check_time, '%H:%M')
                                except (ValueError, TypeError):
                                    pass
                            i += 2
                        else:
                            # Open session - calculate to current time
                            try:
                                now_time = datetime.now().time()
                                end_time_obj = datetime.combine(datetime.today(), now_time)
                                start_time_obj = datetime.combine(datetime.today(), start_time_obj.time())
                            except (ValueError, TypeError):
                                end_time_obj = None
                            i += 1
                        
                        if end_time_obj:
                            if end_time_obj < start_time_obj:
                                end_time_obj += timedelta(days=1)
                            total_working_minutes += (end_time_obj - start_time_obj).total_seconds() / 60
                    else:
                        i += 1
                else:
                    i += 1
            
            # Get break records for this date
            break_records = BreakTimer.objects.filter(
                employee=employee,
                date=check_date
            )
            
            # Calculate total break time and get break details
            total_break_minutes = 0
            break_details = []
            total_extend_count = 0
            for br in break_records:
                total_extend_count += br.extend_count
                if br.break_end_time:
                    try:
                        try:
                            start = datetime.strptime(br.break_start_time, '%H:%M:%S')
                        except ValueError:
                            start = datetime.strptime(br.break_start_time, '%H:%M')
                        try:
                            end = datetime.strptime(br.break_end_time, '%H:%M:%S')
                        except ValueError:
                            end = datetime.strptime(br.break_end_time, '%H:%M')
                        duration = (end - start).total_seconds() / 60
                        total_break_minutes += duration
                        break_details.append(f"{br.get_break_type_display()}")
                    except ValueError:
                        pass
                else:
                    break_details.append(f"{br.get_break_type_display()} (Ongoing)")
            
            # Status is already determined in the main view
            status = data['status']
            
            actual_working_minutes = max(0, total_working_minutes - total_break_minutes)
            working_hours = f"{int(actual_working_minutes // 60)}h {int(actual_working_minutes % 60)}m" if total_working_minutes > 0 else 'N/A'
            break_time = f"{int(total_break_minutes // 60)}h {int(total_break_minutes % 60)}m" if total_break_minutes > 0 else 'N/A'
            breaks_text = ', '.join(break_details) if break_details else 'N/A'
            
            row = [
                data['si_no'],
                check_date,
                employee.employeeId,
                employee.employee_name or "N/A",
                check_in_time or "N/A",
                check_out_time or "N/A",
                check_in_location,
                check_out_location,
                status,
                break_time,
                breaks_text,
                total_extend_count,
                working_hours
            ]
            ws.append(row)
            
            # Apply styling to data rows
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Column width adjustment
        column_widths = [8, 12, 15, 20, 15, 15, 20, 20, 12, 15, 20, 12, 15]  # Custom widths for each column
        for i, width in enumerate(column_widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=attendance_export_{timezone.localdate()}.xlsx'
        wb.save(response)
        return response
    

class EmployeeAttendanceDetailView(LoginRequiredMixin,View):
    def get(self, request, pk):
        employee = get_object_or_404(
            Employee, 
            id=pk, 
            is_superuser=False,
            is_staff=False,
            is_deleted=False
        )
        
        date_filter = request.GET.get('date')
        if date_filter:
            target_date_str = date_filter
            try:
                # Validate format but use string for query
                target_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            except ValueError:
                target_date = timezone.localdate()
                target_date_str = target_date.isoformat()
        else:
            target_date = timezone.localdate()
            target_date_str = target_date.isoformat()
        
        # Check date is CharField
        attendance_records = AttendanceCheck.objects.filter(
            employee=employee,
            check_date=target_date_str
        ).order_by('check_time')
        
        # BreakTimer date is CharField
        break_records = BreakTimer.objects.filter(
            employee=employee,
            date=target_date_str
        ).prefetch_related('extended_breaks').order_by('break_start_time')
        
        total_break_minutes = 0
        total_extend_count = 0
        for break_record in break_records:
            total_extend_count += break_record.extend_count
            if break_record.break_end_time:
                try:
                    try:
                        start_time = datetime.strptime(break_record.break_start_time, '%H:%M:%S')
                    except ValueError:
                        try:
                            start_time = datetime.strptime(break_record.break_start_time, '%H:%M')
                        except ValueError:
                            continue

                    try:
                        end_time = datetime.strptime(break_record.break_end_time, '%H:%M:%S')
                    except ValueError:
                        try:
                            end_time = datetime.strptime(break_record.break_end_time, '%H:%M')
                        except ValueError:
                            continue

                    break_duration = (end_time - start_time).total_seconds() / 60
                    total_break_minutes += break_duration
                except ValueError:
                    pass
        
        total_break_time = f"{int(total_break_minutes // 60)}h {int(total_break_minutes % 60)}m"
        
        check_in = attendance_records.filter(check_type='in').first()
        check_out = attendance_records.filter(check_type='out').last()
        if check_in and check_out:
            # Calculate Total Working Hours from Multiple Sessions
            total_working_minutes = 0
            
            # Filter valid records for calculation
            # We need to process them in chronological order
            attendance_list = list(attendance_records)
            
            i = 0
            while i < len(attendance_list):
                current_record = attendance_list[i]
                
                if current_record.check_type == 'in':
                    start_time = None
                    try:
                        start_time = datetime.strptime(current_record.check_time, '%H:%M:%S')
                    except ValueError:
                        try:
                            start_time = datetime.strptime(current_record.check_time, '%H:%M')
                        except ValueError:
                            pass
                            
                    if start_time:
                        # Find corresponding check-out
                        end_time = None
                        if i + 1 < len(attendance_list) and attendance_list[i+1].check_type == 'out':
                            next_record = attendance_list[i+1]
                            try:
                                end_time = datetime.strptime(next_record.check_time, '%H:%M:%S')
                            except ValueError:
                                try:
                                    end_time = datetime.strptime(next_record.check_time, '%H:%M')
                                except ValueError:
                                    pass
                            i += 2 # Move past this pair
                        else:
                            # No checkout found (Open session)
                            # Handle auto-checkout logic for display purposes
                            # If date is today and time is past 10 PM (22:00), or date is in past
                            is_today = (target_date == timezone.localdate())
                            current_now = timezone.now()
                            
                            cutoff_time = datetime.strptime("22:00:00", "%H:%M:%S").time()
                            
                            should_auto_close = False
                            if target_date < timezone.localdate():
                                should_auto_close = True
                            elif is_today and current_now.time() > cutoff_time:
                                should_auto_close = True
                                
                            if should_auto_close:
                                end_time = start_time.replace(hour=22, minute=0, second=0)
                                if start_time > end_time:
                                     end_time = start_time 
                            
                            i += 1 

                        if end_time:
                            if end_time < start_time:
                                end_time += timedelta(days=1)
                                
                            duration = (end_time - start_time).total_seconds() / 60
                            total_working_minutes += duration

                else:
                    # Starting with 'out' record (orphan), skip
                    i += 1

            # Subtract total break time
            actual_working_minutes = total_working_minutes - total_break_minutes
            if actual_working_minutes < 0: actual_working_minutes = 0
            
            working_hours = f"{int(actual_working_minutes // 60)}h {int(actual_working_minutes % 60)}m" if total_working_minutes > 0 else "N/A"
        else:
            working_hours = "N/A"
        
        if request.GET.get('export') == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from django.http import HttpResponse
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Details"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Set column widths
            column_widths = {
                'A': 15,  # Employee ID
                'B': 20,  # Employee Name
                'C': 12,  # Check Type
                'D': 12,  # Date
                'E': 10,  # Time
                'F': 30,  # Break Records
                'G': 25   # Location
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Headers
            headers = ['Employee ID', 'Employee Name', 'Check Type', 'Date', 'Time', 'Break Records', 'Location']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Prepare break records data
            break_data = {}
            for br in break_records:
                break_info = f"{br.get_break_type_display()}: {br.break_start_time} - {br.break_end_time or 'Ongoing'}"
                if br.extend_count > 0:
                    break_info += f" (Extended {br.extend_count} times)"
                break_data[br.break_start_time] = break_info
            
            # Combine attendance and break data chronologically
            all_records = []
            
            # Add attendance records
            for record in attendance_records:
                # Helper function to parse time strings in different formats
                def parse_time_string(time_str):
                    """Parse time string that can be in format 'HH:MM' or 'HH:MM:SS'"""
                    if not time_str:
                        return None
                    try:
                        return datetime.strptime(time_str, '%H:%M:%S')
                    except ValueError:
                        try:
                            return datetime.strptime(time_str, '%H:%M')
                        except ValueError:
                            return None

                break_info = ""
                # Find break records around this time
                for break_start, break_info_text in break_data.items():
                    try:
                        record_time = parse_time_string(record.check_time)
                        break_start_time = parse_time_string(break_start)
                        if record_time and break_start_time and abs((record_time - break_start_time).total_seconds()) < 3600:  # Within 1 hour
                            break_info = break_info_text
                            break
                    except (ValueError, TypeError):
                        # Skip if time parsing fails
                        pass
                
                all_records.append({
                    'check_type': 'Check In' if record.check_type == 'in' else 'Check Out',
                    'time': record.check_time,
                    'location': record.location or '',
                    'break_records': break_info
                })
            
            # Write data rows
            row_num = 2
            for record in all_records:
                ws.cell(row=row_num, column=1, value=employee.employeeId).border = border
                ws.cell(row=row_num, column=2, value=employee.employee_name or '').border = border
                ws.cell(row=row_num, column=3, value=record['check_type']).border = border
                ws.cell(row=row_num, column=4, value=target_date_str).border = border
                ws.cell(row=row_num, column=5, value=record['time']).border = border
                ws.cell(row=row_num, column=6, value=record['break_records']).border = border
                ws.cell(row=row_num, column=7, value=record['location']).border = border
                row_num += 1
            
            # If no attendance records, still show employee info
            if not all_records:
                ws.cell(row=row_num, column=1, value=employee.employeeId).border = border
                ws.cell(row=row_num, column=2, value=employee.employee_name or '').border = border
                ws.cell(row=row_num, column=3, value='No records').border = border
                ws.cell(row=row_num, column=4, value=target_date_str).border = border
                ws.cell(row=row_num, column=5, value='-').border = border
                ws.cell(row=row_num, column=6, value='-').border = border
                ws.cell(row=row_num, column=7, value='-').border = border
            
            # Prepare response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="attendance_detail_{employee.employeeId}_{target_date_str}.xlsx"'
            
            wb.save(response)
            return response

        context = {
            'employee': employee,
            'selected_date': target_date, # Date object for template display/logic
            'attendance_records': attendance_records,
            'break_records': break_records,
            'total_break_time': total_break_time,
            'check_in': check_in,
            'check_out': check_out,
            'working_hours': working_hours,
            'total_extend_count': total_extend_count,
        }
        
        return render(request, 'employee_attendance_detail.html', context)


class ReportIssueListView(LoginRequiredMixin, View):
    def get(self, request):
        reports_list = ReportIssue.objects.select_related('employee').prefetch_related('media_files').all()
        
        # Pagination
        paginator = Paginator(reports_list, 10)
        page_number = request.GET.get('page')
        reports = paginator.get_page(page_number)
        
        context = {
            'reports': reports,
        }
        return render(request, 'report_list.html', context)


class CompanyAnnouncementListView(LoginRequiredMixin, ListView):
    login_url = '/admin-login/'
    model = CompanyAnnouncement
    template_name = 'company_announcements.html'
    context_object_name = 'announcements'
    
    def get_queryset(self):
        return CompanyAnnouncement.objects.filter(is_active=True)


class CompanyAnnouncementCreateView(LoginRequiredMixin, CreateView):
    login_url = '/admin-login/'
    model = CompanyAnnouncement
    fields = ['heading', 'description', 'date', 'is_active']
    success_url = reverse_lazy('company-announcements-list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Announcement created successfully!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return redirect('company-announcements-list')


class CompanyAnnouncementDeleteView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def delete(self, request, pk):
        try:
            announcement = CompanyAnnouncement.objects.get(pk=pk)
            announcement_heading = announcement.heading
            announcement.delete()
            
            # Return JSON for AJAX requests
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': f'Announcement "{announcement_heading}" deleted successfully!'})
            
            messages.success(request, f'Announcement "{announcement_heading}" deleted successfully!')
            return redirect('company-announcements-list')
            
        except CompanyAnnouncement.DoesNotExist:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': 'Announcement not found.'}, status=404)
            messages.error(request, 'Announcement not found.')
            return redirect('company-announcements-list')
        except Exception as e:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': f'Error deleting announcement: {str(e)}'}, status=500)
            messages.error(request, f'Error deleting announcement: {str(e)}')
            return redirect('company-announcements-list')



class AdminDashboard(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request):
        today_dubai = timezone.localtime()
        today_date = timezone.localdate()
        today_str = today_date.isoformat() # Convert to string for charfield comparison if needed, or use specific format saved
        # Note: AttendanceCheck.check_date is CharField. We need to match the format it saves. 
        # Typically YYYY-MM-DD or similar. Let's assume common formats or check 'today_str'.
        
        current_month = today_date.month
        current_year = today_date.year
        
        first_day_of_month = date(current_year, current_month, 1)
        last_day_of_month = date(current_year, current_month + 1, 1) - timedelta(days=1) if current_month < 12 else date(current_year, 12, 31)
        
        start_of_week = today_date - timedelta(days=today_date.weekday())
        
        # Prepare filters for Admin restriction
        emp_filters = {}
        related_filters = {}
        delivery_task_filters = {}
        
        if getattr(request.user, 'role', '') == 'admin' and not request.user.is_superuser:
            if request.user.company:
                emp_filters['company'] = request.user.company
                related_filters['employee__company'] = request.user.company
                delivery_task_filters['task__employee__company'] = request.user.company
            else:
                # If admin has no company assigned, show nothing
                # Using a condition that is always False/Empty
                emp_filters['pk__in'] = []
                related_filters['pk__in'] = []
                delivery_task_filters['pk__in'] = []
        
        total_employees = Employee.objects.filter(
            is_superuser=False, 
            is_deleted=False,
            role='employee',
            **emp_filters
        ).count()
        
        active_employees = Employee.objects.filter(
            is_superuser=False,
            is_active=True,
            is_deleted=False,
            role='employee',
            **emp_filters
        ).count()
        
        inactive_employees = Employee.objects.filter(
            is_superuser=False,
            is_active=False,
            is_deleted=False,
            role='employee',
            **emp_filters
        ).count()

        total_admins = Employee.objects.filter(
            is_superuser=False,
            is_deleted=False,
            role='admin',
            **emp_filters
        ).count()

        active_admins = Employee.objects.filter(
            is_superuser=False,
            is_active=True,
            is_deleted=False,
            role='admin',
            **emp_filters
        ).count()

        inactive_admins = Employee.objects.filter(
            is_superuser=False,
            is_active=False,
            is_deleted=False,
            role='admin',
            **emp_filters
        ).count()

        total_active_staff = active_employees + active_admins
        
        # 2. Attendance Statistics (Today)
        # Assuming check_date is stored as string "YYYY-MM-DD" or similar.
        # We'll try to match exact string first.
        # If check_date format varies, we might need a regex or more complex logic.
        # Let's try matching with standard ISO format string first.
        possible_date_strs = [
            today_date.strftime('%Y-%m-%d'),
            today_date.strftime('%d-%m-%Y'),
            today_date.strftime('%d/%m/%Y'),
        ]
        
        today_attendance = AttendanceCheck.objects.filter(
            check_date__in=possible_date_strs,
            employee__is_superuser=False,
            employee__is_deleted=False,
            employee__role__in=['employee', 'admin'],
            **related_filters
        )
        
        check_in_count = today_attendance.filter(check_type='in').count()
        check_out_count = today_attendance.filter(check_type='out').count()
        
        employees_present_ids = today_attendance.filter(
            check_type='in'
        ).values_list('employee_id', flat=True).distinct()
        
        employees_present = len(set(employees_present_ids))
        
        # 3. Leave Statistics
        pending_leaves = Leave.objects.filter(status='pending', **related_filters).count()
        
        all_approved_leaves = Leave.objects.filter(status='approved', **related_filters)
        approved_leaves_month = 0
        active_leaves_today = 0
        
        for leave in all_approved_leaves:
            try:
                # Use helper to parse dates
                s_date = self.parse_date_string(leave.start_date)
                e_date = self.parse_date_string(leave.end_date)
                
                if s_date:
                    if s_date.month == current_month and s_date.year == current_year:
                        approved_leaves_month += 1
                
                if s_date and e_date:
                    if s_date <= today_date <= e_date:
                        active_leaves_today += 1
            except:
                continue
        
        # 4. Task Statistics
        base_tasks = Task.objects.filter(
            employee__is_superuser=False,
            employee__is_deleted=False,
            employee__role__in=['employee', 'admin'],
            **related_filters
        )
        
        total_tasks = base_tasks.count()
        
        completed_tasks = base_tasks.filter(status__in=['completed', 'delivered', 'returned']).count()
        in_progress_tasks = base_tasks.filter(status='in_progress').count()
        
        # Urgent tasks - Task model does NOT have due_date. Only DeliveryTask does.
        # We can sum up urgencies from sub-models if needed, or just skip global urgency filter for now
        # to avoid crashing. Or check specific children.
        # Let's count DeliveryTasks due today.
        urgent_tasks = DeliveryTask.objects.filter(
            task__status__in=['not_started', 'in_progress', 'on_hold'],
            due_date__lte=today_date,
            **delivery_task_filters
        ).count()
        
        # 5. Today's Task Distribution
        # Use created_at instead of task_assign_time
        today_tasks_by_type = base_tasks.filter(
            created_at__date=today_date
        ).values('task_type').annotate(count=Count('id'))
        
        # 6. Attendance Rate (This Week)
        week_attendance = {}
        for i in range(7):
            day = start_of_week + timedelta(days=i)
            day_strs = [
                day.strftime('%Y-%m-%d'),
                day.strftime('%d-%m-%Y'),
                day.strftime('%d/%m/%Y'),
            ]
            
            day_employee_ids = AttendanceCheck.objects.filter(
                check_date__in=day_strs,
                employee__is_superuser=False,
                employee__is_deleted=False,
                employee__role__in=['employee', 'admin'],
                check_type='in',
                **related_filters
            ).values_list('employee_id', flat=True).distinct()
            
            week_attendance[day.strftime('%a')] = len(set(day_employee_ids))
        
        # 7. Monthly Leave Trend (Already uses all_approved_leaves which is filtered)
        monthly_leaves = []
        for i in range(5, -1, -1):
            month_date_iter = today_date - timedelta(days=30*i)
            # Simple approximation for month iteration
            m_start = date(month_date_iter.year, month_date_iter.month, 1)
            next_month = month_date_iter.month + 1 if month_date_iter.month < 12 else 1
            next_year = month_date_iter.year if month_date_iter.month < 12 else month_date_iter.year + 1
            m_end = date(next_year, next_month, 1) - timedelta(days=1)
            
            leaves_count = 0
            for leave in all_approved_leaves:
                try:
                    s_date = self.parse_date_string(leave.start_date)
                    if s_date and m_start <= s_date <= m_end:
                        leaves_count += 1
                except:
                    continue
            
            monthly_leaves.append({
                'month': m_start.strftime('%b'),
                'count': leaves_count
            })
        
        # 8. Top Performing Employees
        top_employees = Employee.objects.filter(
            is_superuser=False,
            is_active=True,
            is_deleted=False,
            role__in=['employee', 'admin'],
            **emp_filters
        ).annotate(
            task_count=Count('tasks', filter=Q(tasks__status='completed'))
        ).order_by('-task_count')[:5]
        
        # 9. Recent Activities
        recent_leaves = Leave.objects.filter(
            status='pending',
            **related_filters
        ).select_related('employee').order_by('-created_at')[:5]
        
        recent_tasks = base_tasks.select_related('employee').order_by('-created_at')[:5]
        
        recent_attendance = AttendanceCheck.objects.filter(
            employee__is_superuser=False,
            employee__is_deleted=False,
            employee__role__in=['employee', 'admin'],
            **related_filters
        ).select_related('employee').order_by('-created_at')[:5]
        
        # 10. Task Completion Rate
        if total_tasks > 0:
            task_completion_rate = round((completed_tasks / total_tasks) * 100, 1)
        else:
            task_completion_rate = 0
        
        # 11. Attendance Percentage
        attendance_percentage = round((employees_present / total_active_staff * 100), 1) if total_active_staff > 0 else 0
        
        # 12. Break Statistics
        # BreakTimer date is also CharField? models.py says: date = models.CharField(max_length=100...)
        possible_today_strs = [
            today_date.strftime('%Y-%m-%d'),
            today_date.strftime('%d-%m-%Y'),
        ]
        today_breaks = BreakTimer.objects.filter(
            date__in=possible_today_strs,
            employee__is_superuser=False,
            employee__is_deleted=False,
            employee__role__in=['employee', 'admin'],
            **related_filters
        ).count()
        
        not_started_tasks = base_tasks.filter(status='not_started').count()
        
        
        # 13. Apps Statistics (Detailed Breakdown)
        apps_stats = []
        
        # Helper to get status counts
        def get_status_counts(queryset):
            return {
                'total': queryset.count(),
                'completed': queryset.filter(status__in=['completed', 'delivered', 'returned']).count(),
                'in_progress': queryset.filter(status='in_progress').count(),
                'not_started': queryset.filter(status='not_started').count(),
                'pending': queryset.filter(status__in=['not_started', 'in_progress', 'on_hold', 'paused']).count()
            }
            
        # Mechanic Tasks
        mechanic_tasks = base_tasks.filter(task_type='mechanic')
        mechanic_counts = get_status_counts(mechanic_tasks)
        apps_stats.append({
            'name': 'Mechanic',
            'icon': 'build',
            'color': 'primary',
            **mechanic_counts
        })
        
        # Delivery Tasks
        delivery_tasks = base_tasks.filter(task_type='delivery')
        delivery_counts = get_status_counts(delivery_tasks)
        apps_stats.append({
            'name': 'Delivery',
            'icon': 'local_shipping',
            'color': 'info',
            **delivery_counts
        })
        
        # DAX Service Tasks
        dax_tasks = base_tasks.filter(task_type='service', service_dax_tasks__isnull=False).distinct()
        dax_counts = get_status_counts(dax_tasks)
        apps_stats.append({
            'name': 'DAX Service',
            'icon': 'cleaning_services',
            'color': 'success',
            **dax_counts
        })
        
        # Advantage Service Tasks
        advantage_tasks = base_tasks.filter(task_type='service', advantage_details__isnull=False).distinct()
        advantage_counts = get_status_counts(advantage_tasks)
        apps_stats.append({
            'name': 'Advantage Service',
            'icon': 'verified',
            'color': 'warning',
            **advantage_counts
        })
        
        # Office Tasks
        office_tasks = base_tasks.filter(task_type='office')
        office_counts = get_status_counts(office_tasks)
        apps_stats.append({
            'name': 'Office',
            'icon': 'apartment',
            'color': 'secondary',
            **office_counts
        })
        
        context = {
            'total_employees': total_employees,
            'active_employees': active_employees,
            'inactive_employees': inactive_employees,
            'total_admins': total_admins,
            'active_admins': active_admins,
            'inactive_admins': inactive_admins,
            'total_active_staff': total_active_staff,
            'check_in_count': check_in_count,
            'check_out_count': check_out_count,
            'employees_present': employees_present,
            'attendance_percentage': attendance_percentage,
            'today_breaks': today_breaks,
            'pending_leaves': pending_leaves,
            'approved_leaves_month': approved_leaves_month,
            'active_leaves_today': active_leaves_today,
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'in_progress_tasks': in_progress_tasks,
            'urgent_tasks': urgent_tasks,
            'not_started_tasks': not_started_tasks,
            'task_completion_rate': task_completion_rate,
            'today_tasks_by_type': today_tasks_by_type,
            'week_attendance': week_attendance,
            'monthly_leaves': monthly_leaves,
            'top_employees': top_employees,
            'recent_leaves': recent_leaves,
            'recent_tasks': recent_tasks,
            'recent_attendance': recent_attendance,
            'apps_stats': apps_stats,
            'today': today_date.strftime('%B %d, %Y'),
            'current_month': today_date.strftime('%B %Y'),
            'first_day_of_month': first_day_of_month.strftime('%b %d'),
            'last_day_of_month': last_day_of_month.strftime('%b %d'),
            'current_time': today_dubai.strftime('%I:%M %p'),
        }
        
        return render(request, 'dashboard.html', context)
    
    def parse_date_string(self, date_str):
        """
        Helper method to parse date string to date object.
        Handles various date string formats.
        """
        if not date_str:
            return None
        
        date_formats = [
            '%Y-%m-%d',  # 2025-12-22
            '%d-%m-%Y',  # 22-12-2025
            '%m/%d/%Y',  # 12/22/2025
            '%d/%m/%Y',  # 22/12/2025
            '%Y/%m/%d',  # 2025/12/22
            '%B %d, %Y', # December 22, 2025
            '%b %d, %Y', # Dec 22, 2025
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except (ValueError, TypeError):
                continue
        
        # If no format matches, try to extract date using regex
        try:
            # Look for patterns like YYYY-MM-DD or DD-MM-YYYY
            import re
            match = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', date_str)
            if match:
                year, month, day = match.groups()
                return date(int(year), int(month), int(day))
            
            match = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', date_str)
            if match:
                day, month, year = match.groups()
                return date(int(year), int(month), int(day))
        except:
            pass
        
        return None
    








# Create your models here.
class ForgotPasswordView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def post(self, request, employee_id):
        try:
            import json
            data = json.loads(request.body)
            new_password = data.get('password')
            
            if not new_password:
                return JsonResponse({'success': False, 'error': 'Password is required'})
                
            employee = Employee.objects.get(id=employee_id, is_superuser=False, is_deleted=False)
            
            # Set the new password
            employee.set_password(new_password)
            employee.save()
            
            # Send email with new password (if email is configured)
            if employee.email:
                try:
                    send_mail(
                        'Password Reset - Your Account',
                        f'Hello {employee.employee_name or employee.employeeId},\n\n'
                        f'Your password has been reset by the administrator. Your new password is: {new_password}\n\n'
                        f'Please login using this password.\n\n'
                        f'Best regards,\nYour Company',
                        settings.DEFAULT_FROM_EMAIL,
                        fail_silently=True,
                    )
                except Exception as e:
                    print(f"Error sending email: {e}")
            
            return JsonResponse({'success': True})
            
        except Employee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
        





















class TaskDashboardView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request):
        # Get filter parameters
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        employee_filter = request.GET.get('employee')
        company_filter = request.GET.get('company')
        status_filter = request.GET.get('status')
        task_type_filter = request.GET.get('task_type')
        priority_filter = request.GET.get('priority')
        
        # Base queryset
        tasks = Task.objects.select_related('employee').filter(
            employee__is_superuser=False,
            employee__is_staff=False,
            employee__is_deleted=False
        )
        
        # Apply filters
        if from_date or to_date:
            try:
                date_q = Q()
                if from_date:
                    start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                    date_q &= (Q(created_at__date__gte=start_date) | Q(updated_at__date__gte=start_date) | Q(delivery_details__due_date__gte=start_date))
                if to_date:
                    end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
                    date_q &= (Q(created_at__date__lte=end_date) | Q(updated_at__date__lte=end_date) | Q(delivery_details__due_date__lte=end_date))
                tasks = tasks.filter(date_q)
            except ValueError:
                pass
        
        if employee_filter:
            tasks = tasks.filter(
                Q(employee__employeeId__icontains=employee_filter) |
                Q(employee__employee_name__icontains=employee_filter)
            )

        if company_filter:
            tasks = tasks.filter(employee__company_id=company_filter)
        
        if status_filter:
            tasks = tasks.filter(status=status_filter)
        
        if task_type_filter:
            tasks = tasks.filter(task_type=task_type_filter)
            
        if priority_filter:
            tasks = tasks.filter(priority=priority_filter)
        
        # Statistics
        total_tasks = tasks.count()
        completed_tasks = tasks.filter(status__in=['completed', 'delivered', 'returned']).count()
        in_progress_tasks = tasks.filter(status='in_progress').count()
        not_started_tasks = tasks.filter(status='not_started').count()
        
        # Tasks by type
        tasks_by_type = tasks.values('task_type').annotate(count=Count('id'))
        
        today = timezone.localdate()
        
        # Tasks in date range (or today's tasks if no date range selected)
        if not from_date and not to_date:
            date_range_tasks = tasks.filter(
                Q(delivery_details__task_assign_datetime__date=today) |
                Q(delivery_details__due_date=today) |
                Q(created_at__date=today)
            ).order_by('-priority', '-created_at')
        else:
            date_range_tasks = tasks.order_by('-priority', '-created_at')
        
        # Urgent tasks (due today or overdue) - mainly applicable for delivery tasks
        urgent_tasks = tasks.filter(
            Q(delivery_details__due_date__lte=today)
        ).exclude(status='completed').order_by('delivery_details__due_date')
        
        context = {
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'in_progress_tasks': in_progress_tasks,
            'not_started_tasks': not_started_tasks,
            'tasks_by_type': tasks_by_type,
            'date_range_tasks': date_range_tasks[:50],  # Limit to 50 for performance
            'urgent_tasks': urgent_tasks[:50],
            'from_date': from_date or '',
            'to_date': to_date or '',
            'current_employee': employee_filter or '',
            'current_company': company_filter or '',
            'current_status': status_filter or '',
            'current_task_type': task_type_filter or '',
            'current_priority': priority_filter or '',
            'companies': Company.objects.filter(is_deleted=False),
            'employees': Employee.objects.filter(is_active=True, is_superuser=False, is_staff=False, is_deleted=False),
            'status_choices': Task.TASK_STATUS_CHOICES,
            'task_type_choices': Task.TASK_TYPE_CHOICES,
            'priority_choices': Task.PRIORITY_CHOICES,
            'today': today,
        }
        
        return render(request, 'task_dashboard.html', context)


class GetEmployeesByCompanyView(LoginRequiredMixin, View):
    """API endpoint to get employees filtered by company"""
    login_url = '/admin-login/'
    
    def get(self, request):
        company_id = request.GET.get('company_id')
        
        employees = Employee.objects.filter(
            is_active=True,
            is_superuser=False,
            is_staff=False,
            is_deleted=False
        )
        
        if company_id:
            employees = employees.filter(company_id=company_id)
        
        employee_list = [
            {
                'employeeId': emp.employeeId,
                'employee_name': emp.employee_name or 'N/A'
            }
            for emp in employees
        ]
        
        return JsonResponse({'employees': employee_list})


class TaskListView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request):
        # Get filter parameters
        date_filter = request.GET.get('date')
        employee_filter = request.GET.get('employee')
        status_filter = request.GET.get('status')
        task_type_filter = request.GET.get('task_type')
        priority_filter = request.GET.get('priority')
        company_filter = request.GET.get('company')
        
        # Base queryset with correct related names
        tasks = Task.objects.select_related('employee', 'employee__company').prefetch_related(
            'delivery_details', 
            'mechanic_details',
            'advantage_details',
            'service_dax_tasks'
        ).filter(
            employee__is_superuser=False,
            employee__is_staff=False,
            employee__is_deleted=False
        )
        
        # Apply filters
        if date_filter:
            try:
                target_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
                # Filter by created_at or updated_at as they exist on all Tasks
                tasks = tasks.filter(
                    Q(created_at__date=target_date) |
                    Q(updated_at__date=target_date)
                )
            except ValueError:
                pass
        
        if company_filter:
            tasks = tasks.filter(employee__company_id=company_filter)

        if employee_filter:
            tasks = tasks.filter(
                Q(employee__employeeId=employee_filter) |
                Q(employee__employee_name__icontains=employee_filter)
            )
        
        if status_filter:
            if status_filter == 'delivered':
                tasks = tasks.filter(status='delivered').exclude(
                    task_type='delivery',
                    delivery_details__status_of_delivery__iexact='rejected'
                )
            elif status_filter == 'returned':
                tasks = tasks.filter(
                    Q(status='returned') | 
                    Q(status='delivered', task_type='delivery', delivery_details__status_of_delivery__iexact='rejected')
                )
            else:
                tasks = tasks.filter(status=status_filter)
        
        if task_type_filter:
            tasks = tasks.filter(task_type=task_type_filter)
            
        if priority_filter:
            tasks = tasks.filter(priority=priority_filter)
        
        # Order by priority and created_at
        tasks = tasks.order_by('-priority', '-created_at')
        
        # Pagination
        paginator = Paginator(tasks, 15)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        from apps.authapp.models import Company
        context = {
            'tasks': page_obj,
            'page_obj': page_obj,
            'current_date': date_filter or '',
            'current_employee': employee_filter or '',
            'current_status': status_filter or '',
            'current_task_type': task_type_filter or '',
            'current_priority': priority_filter or '',
            'current_company': company_filter or '',
            'employees': Employee.objects.filter(is_active=True, is_superuser=False, is_staff=False, is_deleted=False),
            'companies': Company.objects.all().order_by('company_name'),
            'status_choices': Task.TASK_STATUS_CHOICES,
            'task_type_choices': Task.TASK_TYPE_CHOICES,
            'priority_choices': Task.PRIORITY_CHOICES,
        }
        
        return render(request, 'task_list.html', context)


class DaxServiceDetailView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request, task_id):
        try:
            task = Task.objects.select_related(
                'employee'
            ).prefetch_related(
                'service_dax_tasks',
                'service_dax_tasks__service_dax_types',
                'service_dax_tasks__completion_images',
                'service_dax_tasks__invoice_images'
            ).get(
                id=task_id,
                task_type='service',
                employee__is_superuser=False,
                employee__is_staff=False,
                employee__is_deleted=False
            )
            
            # Get the first DAX service for this task
            dax_service = task.service_dax_tasks.first()
            
            context = {
                'task': task,
                'dax_service': dax_service,
                'status_choices': Task.TASK_STATUS_CHOICES,
                'priority_choices': Task.PRIORITY_CHOICES,
                'invoice_status_choices': ServiceTaskDax.INVOICE_STATUS_CHOICES,
            }
            return render(request, 'dax_service_detail.html', context)
            
        except Task.DoesNotExist:
            messages.error(request, 'Task not found')
            return redirect('dashboard-dax-service-list')

class AdvantageServiceDetailView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request, task_id):
        try:
            task = Task.objects.select_related(
                'employee',
                'advantage_details',
                'advantage_details__plu'
            ).prefetch_related(
                'advantage_details__images',
                'advantage_details__completion_images'
            ).get(
                id=task_id,
                task_type='service',
                employee__is_superuser=False,
                employee__is_staff=False,
                employee__is_deleted=False
            )
            
            advantage_service = task.advantage_details
            
            context = {
                'task': task,
                'advantage_service': advantage_service,
                'status_choices': Task.TASK_STATUS_CHOICES,
                'priority_choices': Task.PRIORITY_CHOICES,
            }
            return render(request, 'advantage_service_detail.html', context)
        except Task.DoesNotExist:
            messages.error(request, 'Task not found')
            return redirect('dashboard-advantage-service-list')








class LeaveListView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request):
        # Get filter parameters from request
        from_date = request.GET.get('from_date', '')
        to_date = request.GET.get('to_date', '')
        employee_id = request.GET.get('employee', '')
        status = request.GET.get('status', '')
        category = request.GET.get('category', '')
        
        # Base queryset
        if request.user.is_superuser or request.user.role == 'super_admin':
            leaves = Leave.objects.all().select_related('employee', 'approved_by')
        elif request.user.role == 'admin':
            leaves = Leave.objects.all().select_related('employee', 'approved_by')
            if request.user.company:
                leaves = leaves.filter(employee__company=request.user.company)
            else:
                 # Admin with no company sees nothing (or maybe their own?) 
                 # Consistent with previous logic: if company is required for admin view
                 leaves = leaves.none() 
        else:
            leaves = Leave.objects.filter(employee=request.user).select_related('employee', 'approved_by')
        
        # Apply Filters (Basic DB filters)
        if employee_id:
            leaves = leaves.filter(employee__employeeId=employee_id)
        if status:
            leaves = leaves.filter(status=status)
        if category:
            leaves = leaves.filter(category=category)
            
        # Order by latest first
        leaves = leaves.order_by('-created_at')

        # Convert queryset to list for Python-side processing of CharField dates
        leaves_list = list(leaves)
        filtered_leaves = []
        
        # Helper to parse date string
        def parse_date_str(date_str):
            if not date_str: return None
            # Tries common formats
            formats = ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y']
            for fmt in formats:
                try:
                    return datetime.strptime(str(date_str).strip(), fmt).date()
                except ValueError:
                    continue
            return None

        # Filter by Date Range (Python side)
        filter_from = parse_date_str(from_date) if from_date else None
        filter_to = parse_date_str(to_date) if to_date else None
        
        today = date.today()
        current_month = today.month
        current_year = today.year
        
        # Monthly counts and stats
        employees_monthly_leaves = {}
        active_count = 0
        upcoming_count = 0
        
        processed_leaves = []
        
        for leave in leaves_list:
            # Parse leave dates
            l_start = parse_date_str(leave.start_date)
            l_end = parse_date_str(leave.end_date)
            
            # Filter Logic
            include = True
            if filter_from and l_start and l_start < filter_from:
                include = False
            if filter_to and l_end and l_end > filter_to:
                include = False
            
            if include:
                processed_leaves.append(leave)
                
                # Stats Logic (Active/Upcoming)
                if leave.status == 'approved' and l_start and l_end:
                    if l_start <= today <= l_end:
                        active_count += 1
                    if l_start > today:
                        upcoming_count += 1
            
            # Monthly Count Logic (Approved only)
            if leave.status == 'approved' and l_start:
                if l_start.month == current_month and l_start.year == current_year:
                    emp_id = leave.employee.employeeId
                    employees_monthly_leaves[emp_id] = employees_monthly_leaves.get(emp_id, 0) + 1

        # Final filtered list
        leaves = processed_leaves
        
        # Check for Excel export
        if request.GET.get('export') == 'excel':
            return self.export_leaves_to_excel(request, leaves)
        
        # Get statistics counts from the filtered list
        total_leaves = len(leaves)
        pending_count = sum(1 for l in leaves if l.status == 'pending')
        approved_count = sum(1 for l in leaves if l.status == 'approved')
        rejected_count = sum(1 for l in leaves if l.status == 'rejected')
        cancelled_count = sum(1 for l in leaves if l.status == 'cancelled')
        
        # Get employees for dropdown
        if request.user.is_superuser or request.user.role == 'super_admin':
            employees = Employee.objects.filter(
                is_superuser=False,
                is_active=True,
                is_deleted=False
            ).order_by('employee_name')
        elif request.user.role == 'admin':
            if request.user.company:
                employees = Employee.objects.filter(
                    company=request.user.company,
                    is_superuser=False,
                    is_active=True,
                    is_deleted=False
                ).order_by('employee_name')
            else:
                employees = Employee.objects.none()
        else:
            employees = Employee.objects.filter(id=request.user.id)
        
        # Prepare context
        context = {
            'leaves': leaves,
            'employees': employees,
            'status_choices': Leave.STATUS_CHOICES,
            'category_choices': Leave.LEAVE_CATEGORY_CHOICES,
            'ticket_eligibility_choices': Leave.TICKET_ELIGIBILITY_CHOICES,
            
            # Filter values
            'selected_employee': employee_id,
            'selected_status': status,
            'selected_category': category,
            'from_date': from_date,
            'to_date': to_date,
            
            # Statistics
            'total_leaves': total_leaves,
            'pending_count': pending_count,
            'approved_count': approved_count,
            'rejected_count': rejected_count,
            'cancelled_count': cancelled_count,
            'active_count': active_count,
            'upcoming_count': upcoming_count,
            
            # Monthly leave counts
            'employees_monthly_leaves': employees_monthly_leaves,
            'current_month': current_month,
            'current_year': current_year,
            
            # User info
            'is_admin': request.user.is_staff or request.user.is_superuser or request.user.role in ['super_admin', 'admin'],
            'is_employee': request.user.role == 'employee',
        }
        
        return render(request, 'leave_list.html', context)
    
    def export_leaves_to_excel(self, request, leaves):
        """Export leaves to Excel with columns: Sl.No, Employee Name, Leave Details, Type, Duration, Status, Applied On, Approved On"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Leave_Applications_{date.today()}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leave Applications'
        
        # Define styles
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = 'LEAVE APPLICATIONS EXPORT'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers - Sl.No first, then Employee Name, including new date columns
        headers = [
            'Sl.No', 'Employee Name', 'Leave Details', 'Type', 'Reason for Leave', 'Duration', 'Status', 'Applied On', 'Action Date'
        ]
        ws.append(headers)
        
        # Apply styles to header row (Row 2)
        for col_num, cell in enumerate(ws[2], 1):
            cell.fill = yellow_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment_center
        
        # Set column widths
        ws.column_dimensions['A'].width = 8   # Sl.No
        ws.column_dimensions['B'].width = 25  # Employee Name
        ws.column_dimensions['C'].width = 30  # Leave Details
        ws.column_dimensions['D'].width = 15  # Type
        ws.column_dimensions['E'].width = 30  # Reason for Leave
        ws.column_dimensions['F'].width = 12  # Duration
        ws.column_dimensions['G'].width = 15  # Status
        ws.column_dimensions['H'].width = 20  # Applied On
        ws.column_dimensions['I'].width = 20  # Action Date
        
        # Add data
        for index, leave in enumerate(leaves, 1):
            employee_name = leave.employee.employee_name or leave.employee.employeeId or 'N/A'
            leave_details = f"{leave.start_date} - {leave.end_date}"
            leave_type = leave.get_category_display() or 'N/A'
            reason_for_leave = leave.reason or 'N/A'
            duration = str(leave.total_days) + ' days' if leave.total_days else 'N/A'
            status = leave.get_status_display() or 'N/A'
            
            # Convert to local timezone before formatting
            applied_on = timezone.localtime(leave.created_at).strftime('%Y-%m-%d %H:%M') if leave.created_at else 'N/A'
            
            # Dynamic action date based on status
            if leave.status == 'approved':
                action_date = timezone.localtime(leave.approved_at).strftime('%Y-%m-%d %H:%M') if leave.approved_at else 'N/A'
            elif leave.status == 'rejected':
                action_date = timezone.localtime(leave.approved_at).strftime('%Y-%m-%d %H:%M') if leave.approved_at else 'N/A'
            else:
                action_date = 'N/A'
            
            row = [
                index,
                employee_name,
                leave_details,
                leave_type,
                reason_for_leave,
                duration,
                status,
                applied_on,
                action_date
            ]
            ws.append(row)
            
            # Set styles for the newly appended row
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = alignment_left
        
        wb.save(response)
        return response


class LeaveDetailView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request, pk):
        try:
            leave = Leave.objects.select_related('employee', 'approved_by').prefetch_related('attachments', 'signatures').get(pk=pk)
            
            # Check permission
            if not (request.user.is_staff or 
                    request.user.is_superuser or 
                    request.user.role in ['super_admin', 'admin'] or 
                    leave.employee == request.user):
                return redirect('leave-list')
            
            # Calculate monthly leave count for this employee using Python filtering
            current_month = datetime.now().month
            current_year = datetime.now().year
            
            # Helper to parse date string
            def parse_date_str(date_str):
                if not date_str: return None
                formats = ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y']
                for fmt in formats:
                    try:
                        return datetime.strptime(str(date_str).strip(), fmt).date()
                    except ValueError:
                        continue
                return None

            # Calculate monthly leave count - count all approved leaves that overlap with current month
            from datetime import date
            month_start = date(current_year, current_month, 1)
            if current_month == 12:
                month_end = date(current_year + 1, 1, 1)
            else:
                month_end = date(current_year, current_month + 1, 1)
            
            monthly_leave_count = 0
            approved_leaves = Leave.objects.filter(
                employee=leave.employee,
                status='approved'
            )

            for l in approved_leaves:
                leave_start = parse_date_str(l.start_date)
                leave_end = parse_date_str(l.end_date)
                
                # Count leave if it overlaps with current month
                # Overlap condition: leave starts before month ends AND ends after month starts
                if leave_start and leave_end:
                    if leave_start < month_end and leave_end >= month_start:
                        monthly_leave_count += 1
                elif leave_start:
                    # If only start date available, check if it falls in current month
                    if leave_start.month == current_month and leave_start.year == current_year:
                        monthly_leave_count += 1
            
            context = {
                'leave': leave,
                'monthly_leave_count': monthly_leave_count,
                'current_month': current_month,
                'current_year': current_year,
                'is_admin': request.user.is_staff or request.user.is_superuser or request.user.role in ['super_admin', 'admin'],
            }
            return render(request, 'leave_detail.html', context)
            
        except Leave.DoesNotExist:
            messages.error(request, 'Leave not found.')
            return redirect('dashboard-leave-list')


class ApproveLeaveView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def post(self, request, pk):
        # Only admins can approve leaves
        if not (request.user.is_staff or request.user.is_superuser or request.user.role in ['super_admin', 'admin']):
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
            messages.error(request, 'Permission denied.')
            return redirect('dashboard-leave-detail', pk=pk)
        
        try:
            leave = Leave.objects.get(pk=pk)
            
            if leave.status != 'pending':
                if request.headers.get('Content-Type') == 'application/json':
                    return JsonResponse({'success': False, 'error': 'Leave is not in pending status.'}, status=400)
                messages.error(request, 'Leave is not in pending status.')
                return redirect('dashboard-leave-detail', pk=pk)
            
            leave.status = 'approved'
            leave.approved_by = request.user
            leave.approved_at = timezone.now()
            leave.save()
            
            # Recalculate leave balance for the month of the approved leave
            from datetime import datetime
            try:
                leave_start_date = datetime.strptime(leave.start_date, '%Y-%m-%d')
                calculate_leave_balance(leave.employee, leave_start_date.year, leave_start_date.month)
            except (ValueError, TypeError):
                # If date parsing fails, use current month
                calculate_leave_balance(leave.employee, timezone.now().year, timezone.now().month)
            
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': 'Leave approved successfully.'})
            
            messages.success(request, 'Leave approved successfully.')
            return redirect('dashboard-leave-detail', pk=pk)
            
        except Leave.DoesNotExist:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': 'Leave not found.'}, status=404)
            messages.error(request, 'Leave not found.')
            return redirect('dashboard-leave-list')
        except Exception as e:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': f'Error approving leave: {str(e)}'}, status=500)
            messages.error(request, f'Error approving leave: {str(e)}')
            return redirect('dashboard-leave-detail', pk=pk)


class RejectLeaveView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def post(self, request, pk):
        # Only admins can reject leaves
        if not (request.user.is_staff or request.user.is_superuser or request.user.role in ['super_admin', 'admin']):
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
            messages.error(request, 'Permission denied.')
            return redirect('dashboard-leave-detail', pk=pk)
        
        try:
            # Parse JSON body if it's an AJAX request
            if request.headers.get('Content-Type') == 'application/json':
                import json
                data = json.loads(request.body)
                rejection_reason = data.get('rejection_reason', '').strip()
            else:
                rejection_reason = request.POST.get('rejection_reason', '').strip()
            
            leave = Leave.objects.get(pk=pk)
            
            if leave.status != 'pending':
                if request.headers.get('Content-Type') == 'application/json':
                    return JsonResponse({'success': False, 'error': 'Leave is not in pending status.'}, status=400)
                messages.error(request, 'Leave is not in pending status.')
                return redirect('dashboard-leave-detail', pk=pk)
            
            if not rejection_reason:
                if request.headers.get('Content-Type') == 'application/json':
                    return JsonResponse({'success': False, 'error': 'Rejection reason is required.'}, status=400)
                messages.error(request, 'Rejection reason is required.')
                return redirect('dashboard-leave-detail', pk=pk)
            
            leave.status = 'rejected'
            leave.rejection_reason = rejection_reason
            leave.approved_by = request.user
            leave.approved_at = timezone.now()
            leave.save()
            
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': 'Leave rejected successfully.'})
            
            messages.success(request, 'Leave rejected successfully.')
            return redirect('dashboard-leave-detail', pk=pk)
            
        except Leave.DoesNotExist:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': 'Leave not found.'}, status=404)
            messages.error(request, 'Leave not found.')
            return redirect('dashboard-leave-list')
        except Exception as e:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': f'Error rejecting leave: {str(e)}'}, status=500)
            messages.error(request, f'Error rejecting leave: {str(e)}')
            return redirect('dashboard-leave-detail', pk=pk)


class CancelLeaveView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def post(self, request, pk):
        try:
            leave = Leave.objects.get(pk=pk)
            
            # Only the employee can cancel their own leave
            if leave.employee != request.user:
                if request.headers.get('Content-Type') == 'application/json':
                    return JsonResponse({'success': False, 'error': 'You can only cancel your own leaves.'}, status=403)
                messages.error(request, 'You can only cancel your own leaves.')
                return redirect('dashboard-leave-detail', pk=pk)
            
            if leave.status != 'pending':
                if request.headers.get('Content-Type') == 'application/json':
                    return JsonResponse({'success': False, 'error': 'Only pending leaves can be cancelled.'}, status=400)
                messages.error(request, 'Only pending leaves can be cancelled.')
                return redirect('dashboard-leave-detail', pk=pk)
            
            leave.status = 'cancelled'
            leave.save()
            
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': True, 'message': 'Leave cancelled successfully.'})
            
            messages.success(request, 'Leave cancelled successfully.')
            return redirect('dashboard-leave-detail', pk=pk)
            
        except Leave.DoesNotExist:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': 'Leave not found.'}, status=404)
            messages.error(request, 'Leave not found.')
            return redirect('dashboard-leave-list')
        except Exception as e:
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({'success': False, 'error': f'Error cancelling leave: {str(e)}'}, status=500)
            messages.error(request, f'Error cancelling leave: {str(e)}')
            return redirect('dashboard-leave-detail', pk=pk)


# Admin Profile Views
class AdminProfileView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request):
        # Only allow admin and super admin
        if not (request.user.is_staff or request.user.is_superuser or request.user.role in ['super_admin', 'admin']):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('dashboard')
        
        context = {
            'user': request.user,
        }
        return render(request, 'admin_profile.html', context)


class AdminProfileUpdateView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def post(self, request):
        # Only allow admin and super admin
        if not (request.user.is_staff or request.user.is_superuser or request.user.role in ['super_admin', 'admin']):
            messages.error(request, 'Access denied. Admin privileges required.')
            return redirect('dashboard')
        
        try:
            user = request.user
            
            # Update basic fields
            user.employee_name = request.POST.get('employee_name', user.employee_name)
            user.email = request.POST.get('email', user.email)
            user.mobile_number = request.POST.get('mobile_number', user.mobile_number)
            user.home_address = request.POST.get('home_address', user.home_address)
            user.nationality = request.POST.get('nationality', user.nationality)
            
            # Update date of birth if provided
            dob = request.POST.get('date_of_birth')
            if dob:
                user.date_of_birth = dob
            
            # Update emergency contact
            user.emergency_contact_name = request.POST.get('emergency_contact_name', user.emergency_contact_name)
            user.emergency_contact_number = request.POST.get('emergency_contact_number', user.emergency_contact_number)
            user.emergency_contact_relation = request.POST.get('emergency_contact_relation', user.emergency_contact_relation)
            
            # Handle profile picture upload
            if 'profile_pic' in request.FILES:
                user.profile_pic = request.FILES['profile_pic']
            
            user.save()
            
            messages.success(request, 'Profile updated successfully!')
            return redirect('admin-profile')
            
        except Exception as e:
            messages.error(request, f'Error updating profile: {str(e)}')
            return redirect('admin-profile')


# ============================================
# Task Type Specific Views
# ============================================

class DaxServiceListView(LoginRequiredMixin, View):
    """View for listing DAX Service tasks"""
    login_url = '/admin-login/'
    
    def get(self, request):
        # Get filter parameters
        status_filter = request.GET.get('status')
        employee_filter = request.GET.get('employee')
        search_query = request.GET.get('search')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
        # Base queryset - filter tasks with ServiceTaskDax records
        tasks = Task.objects.filter(
            task_type='service',
            service_dax_tasks__isnull=False,
            employee__is_superuser=False,
            employee__is_staff=False,
            employee__is_deleted=False
        ).select_related('employee').prefetch_related('service_dax_tasks').distinct()
        
        # Apply filters
        if status_filter:
            tasks = tasks.filter(status=status_filter)
        
        if employee_filter:
            tasks = tasks.filter(
                Q(employee__employeeId__icontains=employee_filter) |
                Q(employee__employee_name__icontains=employee_filter)
            )
        
        if search_query:
            tasks = tasks.filter(
                Q(employee__employee_name__icontains=search_query) |
                Q(employee__employeeId__icontains=search_query) |
                Q(service_dax_tasks__chassis_no__icontains=search_query) |
                Q(service_dax_tasks__vehicle_model__icontains=search_query)
            )
        
        # Apply date range filter
        if from_date:
            tasks = tasks.filter(created_at__date__gte=from_date)
        if to_date:
            tasks = tasks.filter(created_at__date__lte=to_date)
        
        # Order by priority and created_at
        tasks = tasks.order_by('-priority', '-created_at')
        
        export_excel = request.GET.get('export') == 'excel'
        if export_excel:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from django.http import HttpResponse

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="DAX_Service_Tasks.xlsx"'

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'DAX Detailing'

            # Define styles
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Title
            ws.merge_cells('A1:Q1')
            title_cell = ws['A1']
            title_cell.value = 'DAX DETAILING - EXCEL REPORT'
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = alignment_center

            # Headers
            headers = [
                'SI NO', 'Employee Name', 'Created At', 'Start Date & Time', 'Completed Date & Time',
                'Detailing Site', 'Service Type', 'Service Sub Type', 'Tinting Film VLT', 'Layers', 'Roll Meter',
                'Consumables', 'Chassis Number', 'Vehicle Model/Type', 'Proof of invoice/PR (Yes/No)', 
                'Completed or Not', 'Final Notes'
            ]
            ws.append(headers)

            # Apply styles to header row (Row 2)
            for col_num, cell in enumerate(ws[2], 1):
                cell.fill = yellow_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = alignment_center
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

            from django.utils import timezone
            
            # Add data
            for index, task in enumerate(tasks, 1):
                dax_service = task.service_dax_tasks.first()
                if not dax_service:
                    continue

                employee_name = task.employee.employee_name if task.employee else 'N/A'
                
                # Format datetimes
                created_dt = timezone.localtime(task.created_at).strftime("%d-%m-%Y %I:%M %p") if task.created_at else 'N/A'
                start_dt = timezone.localtime(dax_service.started_at).strftime("%d-%m-%Y %I:%M %p") if dax_service.started_at else 'N/A'
                completed_dt = timezone.localtime(dax_service.completed_at).strftime("%d-%m-%Y %I:%M %p") if dax_service.completed_at else 'N/A'

                detailing_site = dax_service.detailing_site or 'N/A'
                remarks = dax_service.remarks or ''
                chassis_number = dax_service.chassis_no or 'N/A'
                vehicle_model = dax_service.vehicle_model or 'N/A'
                proof = 'Yes' if dax_service.invoice_status else 'No'
                completed = 'Completed' if task.status == 'completed' else 'Not Completed'
                final_notes = dax_service.completion_remarks or ''
                
                # Fetch service types
                service_types = dax_service.service_dax_types.all()
                if not service_types:
                    st_type = 'N/A'
                    st_sub = 'N/A'
                    st_level = 'N/A'
                    st_layers = 'N/A'
                    st_roll = 'N/A'
                else:
                    st_type = '\n'.join([st.service_type or 'N/A' for st in service_types])
                    st_sub = '\n'.join([st.service_sub_type or 'N/A' for st in service_types])
                    st_level = '\n'.join([st.level or 'N/A' for st in service_types])
                    st_layers = '\n'.join([st.layers or 'N/A' for st in service_types])
                    st_roll = '\n'.join([st.roll_meter or 'N/A' for st in service_types])

                row = [
                    index, employee_name, created_dt, start_dt, completed_dt,
                    detailing_site, st_type, st_sub, st_level, st_layers, st_roll,
                    remarks, chassis_number, vehicle_model, proof, completed, final_notes
                ]
                ws.append(row)
                
                # Set styles for the newly appended row
                for cell in ws[ws.max_row]:
                    cell.border = border
                    cell.alignment = alignment_left

            wb.save(response)
            return response

        # Pagination
        paginator = Paginator(tasks, 15)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'tasks': page_obj,
            'page_obj': page_obj,
            'current_status': status_filter or '',
            'current_employee': employee_filter or '',
            'search_query': search_query or '',
            'from_date': from_date or '',
            'to_date': to_date or '',
            'employees': Employee.objects.filter(is_active=True, is_superuser=False, is_staff=False, is_deleted=False, employee_type='service', company__company_name='dax'),
            'status_choices': Task.TASK_STATUS_CHOICES,
            'priority_choices': Task.PRIORITY_CHOICES,
        }
        
        return render(request, 'dax_service_list.html', context)



class AdvantageServiceListView(LoginRequiredMixin, View):
    """View for listing ADVANTAGE Service tasks"""
    login_url = '/admin-login/'
    
    def get(self, request):
        # Get filter parameters
        status_filter = request.GET.get('status')
        employee_filter = request.GET.get('employee')
        company_filter = request.GET.get('company')
        search_query = request.GET.get('search')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
        # Base queryset - filter tasks with ServiceAdvantage records
        tasks = Task.objects.filter(
            task_type='service',
            advantage_details__isnull=False,
            employee__is_superuser=False,
            employee__is_staff=False,
            employee__is_deleted=False
        ).select_related('employee', 'advantage_details', 'advantage_details__plu')
        
        # Apply company filter
        if company_filter:
            tasks = tasks.filter(employee__company_id=company_filter)
        
        # Apply filters
        if status_filter:
            tasks = tasks.filter(status=status_filter)
        
        if employee_filter:
            tasks = tasks.filter(
                Q(employee__employeeId__icontains=employee_filter) |
                Q(employee__employee_name__icontains=employee_filter)
            )
        
        if search_query:
            tasks = tasks.filter(
                Q(employee__employee_name__icontains=search_query) |
                Q(employee__employeeId__icontains=search_query) |
                Q(advantage_details__chassis_number__icontains=search_query) |
                Q(advantage_details__detailing_site__icontains=search_query)
            )
        
        # Apply date range filter
        if from_date:
            tasks = tasks.filter(created_at__date__gte=from_date)
        if to_date:
            tasks = tasks.filter(created_at__date__lte=to_date)
        
        # Order by priority and created_at
        tasks = tasks.order_by('-priority', '-created_at')
        
        export_excel = request.GET.get('export') == 'excel'
        if export_excel:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from django.http import HttpResponse

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="ADVANTAGE_Service_Tasks.xlsx"'

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'ADVANTAGE Detailing'

            # Define styles
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Title - Expanded to K1 to cover all 11 columns
            ws.merge_cells('A1:K1')
            title_cell = ws['A1']
            title_cell.value = 'ADVANTAGE DETAILING SITE - EXCEL REPORT'
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = alignment_center

            # Updated Headers with SI NO, Consumables and other notes and Created At
            headers = [
                'SI NO', 'Employee Name', 'Created At', 'Detailing Site', 'Material', 'Service Description', 
                'Category', 'Chassis Number', 'Start Date & Time', 'Completed Date & Time', 'Consumables and other notes'
            ]
            ws.append(headers)

            # Apply styles to header row (Row 2)
            for col_num, cell in enumerate(ws[2], 1):
                cell.fill = yellow_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = alignment_center
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

            # Add data
            for index, task in enumerate(tasks, 1):
                adv_service = getattr(task, 'advantage_details', None)
                if not adv_service:
                    continue
                
                # Fetch Employee Name (Fallback to ID if name is missing)
                if task.employee:
                    employee_name = task.employee.employee_name or task.employee.employeeId
                else:
                    employee_name = 'N/A'

                detailing_site = adv_service.detailing_site or 'N/A'
                chassis_number = adv_service.chassis_number or 'N/A'
                completion_remarks = adv_service.completion_remarks or 'N/A'
                
                plu_obj = adv_service.plu
                if plu_obj:
                    material = plu_obj.plu or 'N/A'
                    category = plu_obj.category or 'N/A'
                    service_desc = plu_obj.sub_service or 'N/A'
                else:
                    material = 'N/A'
                    category = 'N/A'
                    service_desc = 'N/A'

                # Format Dates using localtime for consistency
                created_date = timezone.localtime(task.created_at).strftime('%d-%m-%Y %I:%M %p') if task.created_at else 'N/A'
                start_date = timezone.localtime(adv_service.started_at).strftime('%d-%m-%Y %I:%M %p') if adv_service.started_at else 'N/A'
                completed_date = timezone.localtime(adv_service.completed_at).strftime('%d-%m-%Y %I:%M %p') if adv_service.completed_at else 'N/A'

                # Updated Row Mapping with SI NO, Completion Remarks and Created At
                row = [
                    index, employee_name, created_date, detailing_site, material, service_desc, 
                    category, chassis_number, start_date, completed_date, completion_remarks
                ]
                ws.append(row)

                # Set styles for the newly appended row
                for cell in ws[ws.max_row]:
                    cell.border = border
                    cell.alignment = alignment_left

            wb.save(response)
            return response

        # Pagination
        paginator = Paginator(tasks, 15)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Filter employees - always show only ADVANTAGE company service employees
        employees = Employee.objects.filter(
            is_active=True, 
            is_superuser=False, 
            is_staff=False, 
            is_deleted=False, 
            employee_type='service',
            company__company_name='advantage'
        ).order_by('employee_name')
        
        context = {
            'tasks': page_obj,
            'page_obj': page_obj,
            'current_status': status_filter or '',
            'current_company': company_filter or '',
            'current_employee': employee_filter or '',
            'search_query': search_query or '',
            'from_date': from_date or '',
            'to_date': to_date or '',
            'companies': Company.objects.filter(is_deleted=False).order_by('company_name'),
            'employees': employees,
            'status_choices': Task.TASK_STATUS_CHOICES,
            'priority_choices': Task.PRIORITY_CHOICES,
        }
        
        return render(request, 'advantage_service_list.html', context)



class DeliveryTaskListView(LoginRequiredMixin, View):
    """View for listing Delivery tasks"""
    login_url = '/admin-login/'
    
    def get(self, request):
        # Get filter parameters
        status_filter = request.GET.get('status')
        employee_filter = request.GET.get('employee')
        company_filter = request.GET.get('company')
        search_query = request.GET.get('search')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
        # Base queryset - filter delivery tasks
        tasks = Task.objects.filter(
            task_type='delivery',
            delivery_details__isnull=False,
            employee__is_superuser=False,
            employee__is_staff=False,
            employee__is_deleted=False
        ).select_related('employee', 'delivery_details')
        
        # Apply company filter
        if company_filter:
            tasks = tasks.filter(employee__company_id=company_filter)
        
        # Apply filters
        if status_filter:
            if status_filter == 'delivered':
                tasks = tasks.filter(status='delivered').exclude(
                    delivery_details__status_of_delivery__iexact='rejected'
                )
            elif status_filter == 'returned':
                tasks = tasks.filter(
                    Q(status='returned') | 
                    Q(status='delivered', delivery_details__status_of_delivery__iexact='rejected')
                )
            else:
                tasks = tasks.filter(status=status_filter)
        
        if employee_filter:
            tasks = tasks.filter(
                Q(employee__employeeId__icontains=employee_filter) |
                Q(employee__employee_name__icontains=employee_filter)
            )
        
        if search_query:
            tasks = tasks.filter(
                Q(employee__employee_name__icontains=search_query) |
                Q(employee__employeeId__icontains=search_query) |
                Q(delivery_details__customer_name__icontains=search_query) |
                Q(delivery_details__customer_phone__icontains=search_query) |
                Q(delivery_details__DeliveryId__icontains=search_query)
            )
        
        # Apply date range filter
        if from_date or to_date:
            try:
                date_q = Q()
                if from_date:
                    start_date = datetime.strptime(from_date, '%Y-%m-%d').date()
                    date_q &= (Q(created_at__date__gte=start_date) | 
                             Q(delivery_details__task_assign_datetime__date__gte=start_date) |
                             Q(delivery_details__due_date__gte=start_date))
                if to_date:
                    end_date = datetime.strptime(to_date, '%Y-%m-%d').date()
                    date_q &= (Q(created_at__date__lte=end_date) | 
                             Q(delivery_details__task_assign_datetime__date__lte=end_date) |
                             Q(delivery_details__due_date__lte=end_date))
                tasks = tasks.filter(date_q)
            except ValueError:
                pass
        
        # Order by priority and created_at
        tasks = tasks.order_by('-priority', '-created_at')
        
        export_excel = request.GET.get('export') == 'excel'
        if export_excel:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from django.http import HttpResponse
            from apps.task.models import DeliveryNote

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Delivery_Tasks.xlsx"'

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'DELIVERY TASKS'

            # Define styles
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Title
            ws.merge_cells('A1:L1')
            title_cell = ws['A1']
            title_cell.value = 'DELIVERY TASKS EXPORT'
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Headers - Sl.No first, then Employee Name
            headers = [
                'Sl.No', 'Employee Name', 'Assigned Date & Time', 'Invoice Number', 'Customer Name', 'Customer Phone', 'Delivery Location', 
                'Task Start Date & Time', 'Task Completed Date & Time', 'Status of Delivery', 'Delivery Notes (Task)', 'Employee Delivery Notes'
            ]
            ws.append(headers)

            # Apply styles to header row (Row 2)
            for col_num, cell in enumerate(ws[2], 1):
                cell.fill = yellow_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = alignment_center
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25
            ws.column_dimensions['A'].width = 8   # Sl.No
            ws.column_dimensions['B'].width = 25  # Employee Name
            ws.column_dimensions['C'].width = 25  # Assigned Date & Time
            ws.column_dimensions['D'].width = 15  # Delivery ID
            ws.column_dimensions['E'].width = 25  # Customer Name
            ws.column_dimensions['F'].width = 15  # Customer Phone
            ws.column_dimensions['G'].width = 30  # Delivery Location
            ws.column_dimensions['H'].width = 25  # Start Date & Time
            ws.column_dimensions['I'].width = 25  # Completed Date & Time
            ws.column_dimensions['J'].width = 20  # Status of Delivery
            ws.column_dimensions['K'].width = 30  # Delivery Notes (Task)
            ws.column_dimensions['L'].width = 30  # Employee Delivery Notes

            # Add data
            for index, task in enumerate(tasks, 1):
                del_details = getattr(task, 'delivery_details', None)
                if not del_details:
                    continue
                
                delivery_id = del_details.DeliveryId or 'N/A'
                customer_name = del_details.customer_name or 'N/A'
                customer_phone = del_details.customer_phone or 'N/A'
                delivery_location = del_details.delivery_location or 'N/A'
                delivery_notes = del_details.delivery_notes or 'N/A'
                status_of_delivery = del_details.status_of_delivery or 'N/A'
                
                employee = task.employee
                employee_name = employee.employee_name if employee else 'N/A'
                
                # Assigned Date & Time
                assigned_datetime = 'N/A'
                if task.created_at:
                    assigned_datetime = timezone.localtime(task.created_at).strftime('%d-%m-%Y %I:%M %p')
                
                start_datetime = 'N/A'
                if del_details.task_start_datetime:
                    start_datetime = timezone.localtime(del_details.task_start_datetime).strftime('%d-%m-%Y %I:%M %p')
                
                completed_datetime = 'N/A'
                if del_details.task_completed_date and del_details.task_completed_time:
                    try:
                        dt = datetime.combine(del_details.task_completed_date, del_details.task_completed_time)
                        completed_datetime = dt.strftime('%d-%m-%Y %I:%M %p')
                    except Exception:
                        completed_datetime = f"{del_details.task_completed_date.strftime('%d-%m-%Y')} {del_details.task_completed_time.strftime('%I:%M %p')}"
                
                if employee:
                    emp_notes = DeliveryNote.objects.filter(delivery_task__task__employee=employee).values_list('note', flat=True)
                    emp_delivery_notes = '\n'.join(emp_notes) if emp_notes else 'N/A'
                else:
                    emp_delivery_notes = 'N/A'

                row = [
                    index, employee_name, assigned_datetime, delivery_id, customer_name, customer_phone, delivery_location, 
                    start_datetime, completed_datetime, status_of_delivery, delivery_notes, emp_delivery_notes
                ]
                ws.append(row)

                # Set styles for the newly appended row
                for cell in ws[ws.max_row]:
                    cell.border = border
                    cell.alignment = alignment_left

            wb.save(response)
            return response
        
        # Pagination
        paginator = Paginator(tasks, 15)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Filter employees based on selected company
        if company_filter:
            employees = Employee.objects.filter(
                is_active=True, 
                is_superuser=False, 
                is_staff=False, 
                is_deleted=False, 
                employee_type='deliver',
                company_id=company_filter
            ).order_by('employee_name')
        else:
            employees = Employee.objects.filter(
                is_active=True, 
                is_superuser=False, 
                is_staff=False, 
                is_deleted=False, 
                employee_type='deliver'
            ).order_by('employee_name')
        
        context = {
            'tasks': page_obj,
            'page_obj': page_obj,
            'current_status': status_filter or '',
            'current_company': company_filter or '',
            'current_employee': employee_filter or '',
            'search_query': search_query or '',
            'from_date': from_date or '',
            'to_date': to_date or '',
            'companies': Company.objects.filter(is_deleted=False).order_by('company_name'),
            'employees': employees,
            'status_choices': Task.TASK_STATUS_CHOICES,
            'priority_choices': Task.PRIORITY_CHOICES,
        }
        
        return render(request, 'delivery_task_list.html', context)


class MechanicTaskListView(LoginRequiredMixin, View):
    """View for listing Mechanic tasks"""
    login_url = '/admin-login/'
    
    def get(self, request):
        # Get filter parameters
        status_filter = request.GET.get('status')
        employee_filter = request.GET.get('employee')
        company_filter = request.GET.get('company')
        search_query = request.GET.get('search')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
        # Base queryset - filter mechanic tasks
        tasks = Task.objects.filter(
            Q(employee__isnull=True) | Q(employee__is_superuser=False, employee__is_staff=False, employee__is_deleted=False),
            task_type='mechanic',
            mechanic_details__isnull=False,
        ).select_related('employee', 'mechanic_details')
        
        # Apply company filter
        if company_filter:
            tasks = tasks.filter(employee__company_id=company_filter)
        
        # Apply date filters
        if from_date:
            tasks = tasks.filter(created_at__date__gte=from_date)
        if to_date:
            tasks = tasks.filter(created_at__date__lte=to_date)
        
        # Apply filters
        if status_filter:
            tasks = tasks.filter(status=status_filter)
        
        if employee_filter:
            tasks = tasks.filter(
                Q(employee__employeeId__icontains=employee_filter) |
                Q(employee__employee_name__icontains=employee_filter)
            )
        
        if search_query:
            tasks = tasks.filter(
                Q(employee__employee_name__icontains=search_query) |
                Q(employee__employeeId__icontains=search_query) |
                Q(mechanic_details__heading__icontains=search_query) |
                Q(mechanic_details__Site_number__icontains=search_query) |
                Q(mechanic_details__Machine_serial_number__icontains=search_query)
            )
        
        # Order by priority and created_at
        tasks = tasks.order_by('-priority', '-created_at')
        
        export_excel = request.GET.get('export') == 'excel'
        if export_excel:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from django.http import HttpResponse

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Mechanic_Tasks.xlsx"'

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'MAINTENANCE TASK REPORT'

            # Define styles
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # Title
            ws.merge_cells('A1:O1')
            title_cell = ws['A1']
            title_cell.value = 'MAINTENANCE TASK REPORT'
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Headers - Sl.No first, split part numbers into three columns
            headers = [
                'Sl.No', 'Employee Name', 'Created At', 'Start Date & Time', 'Completed Date & Time', 'SITE', 'BAY No', 'Machine Type', 'Machine Number', 'Job Description', 'Spare Parts Details', 'PARTNUMBER', 'ITEM', 'QUANTITIES', 'Remarks'
            ]
            ws.append(headers)

            # Apply styles to header row (Row 2)
            for col_num, cell in enumerate(ws[2], 1):
                cell.fill = yellow_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = alignment_center
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
            ws.column_dimensions['A'].width = 8   # Sl.No
            ws.column_dimensions['B'].width = 25  # Employee Name
            ws.column_dimensions['C'].width = 25  # Created At
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['L'].width = 20  # PARTNUMBER
            ws.column_dimensions['M'].width = 25  # ITEM
            ws.column_dimensions['N'].width = 15  # QUANTITIES

            from django.utils import timezone

            # Add data
            for index, task in enumerate(tasks, 1):
                mech_details = getattr(task, 'mechanic_details', None)
                if not mech_details:
                    continue
                
                employee_name = task.employee.employee_name if task.employee else 'N/A'
                
                # Format Dates using localtime and consistent format
                created_at_str = timezone.localtime(task.created_at).strftime('%b. %d, %Y, %I:%M %p') if task.created_at else 'N/A'
                started_at_str = timezone.localtime(mech_details.started_at).strftime('%b. %d, %Y, %I:%M %p') if mech_details.started_at else 'N/A'
                completed_at_str = timezone.localtime(mech_details.completed_at).strftime('%b. %d, %Y, %I:%M %p') if mech_details.completed_at else 'N/A'
                
                site = mech_details.Site_number or 'N/A'
                bay_no = mech_details.bay_number or 'N/A'
                machine_type = mech_details.Machine_type or 'N/A'
                machine_number = mech_details.Machine_serial_number or 'N/A'
                job_desc = mech_details.job_description or 'N/A'
                spare_parts = mech_details.spare_part_details or 'N/A'
                
                # Get all part items for this mechanic task
                part_items = mech_details.part_items.all()
                remarks = mech_details.completion_remarks or 'N/A'

                # Build separate lists for each column
                if part_items:
                    part_numbers_list = []
                    item_names_list = []
                    quantities_list = []
                    for item in part_items:
                        part_num = item.part_number.part_number if item.part_number else 'N/A'
                        item_name = item.part_number.item if item.part_number and item.part_number.item else 'N/A'
                        qty = str(item.quantity) if item.quantity else '0'
                        part_numbers_list.append(part_num)
                        item_names_list.append(item_name)
                        quantities_list.append(qty)
                    part_numbers_col = '\n'.join(part_numbers_list)
                    items_col = '\n'.join(item_names_list)
                    quantities_col = '\n'.join(quantities_list)
                else:
                    part_numbers_col = 'N/A'
                    items_col = 'N/A'
                    quantities_col = 'N/A'

                # Create single row per task
                row = [
                    index, employee_name, created_at_str, started_at_str, completed_at_str, 
                    site, bay_no, machine_type, machine_number, job_desc, 
                    spare_parts, part_numbers_col, items_col, quantities_col, remarks
                ]
                ws.append(row)

                # Set styles for the newly appended row
                for cell in ws[ws.max_row]:
                    cell.border = border
                    cell.alignment = alignment_left

            wb.save(response)
            return response
        
        # Pagination
        paginator = Paginator(tasks, 15)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'tasks': page_obj,
            'page_obj': page_obj,
            'current_status': status_filter or '',
            'current_company': company_filter or '',
            'current_employee': employee_filter or '',
            'search_query': search_query or '',
            'companies': Company.objects.filter(is_deleted=False).order_by('company_name'),
            'employees': Employee.objects.filter(is_active=True, is_superuser=False, is_staff=False, is_deleted=False, employee_type='mechanic'),
            'status_choices': Task.TASK_STATUS_CHOICES,
            'priority_choices': Task.PRIORITY_CHOICES,
        }
        
        return render(request, 'mechanic_task_list.html', context)


from apps.task.models import Task, DeliveryTask, DeliveryTaskImage

class CreateDeliveryTaskView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request):
        context = {
            'employees': Employee.objects.filter(is_active=True, is_superuser=False, is_staff=False, is_deleted=False, employee_type='deliver'),
            'status_choices': Task.TASK_STATUS_CHOICES,
            'priority_choices': Task.PRIORITY_CHOICES,
        }
        return render(request, 'create_delivery_task.html', context)
    
    def post(self, request):
        try:
            # 1. Create Base Task
            task = Task(
                employee_id=request.POST.get('employee'),
                task_type='delivery',
                status=request.POST.get('status', 'not_started'),
                priority=request.POST.get('priority', 'medium'),
                icon_type='delivery'
            )
            task.save()
            
            # 2. Extract Delivery Specific Fields
            # Handle date/time fields which might be empty
            task_assign_datetime = request.POST.get('task_assign_datetime') or None
            task_start_datetime = request.POST.get('task_start_datetime') or None
            
            task_completed_date = request.POST.get('task_completed_date') or None
            task_completed_time = request.POST.get('task_completed_time') or None
            
            due_date = request.POST.get('due_date') or None
            due_time = request.POST.get('due_time') or None

            # Create DeliveryTask
            delivery_task = DeliveryTask.objects.create(
                task=task,
                DeliveryId=request.POST.get('DeliveryId'),
                customer_name=request.POST.get('customer_name'),
                customer_phone=request.POST.get('customer_phone'),
                delivery_location=request.POST.get('delivery_location'),
                delivery_notes=request.POST.get('delivery_notes'),
                task_assign_datetime=task_assign_datetime,
                task_start_datetime=task_start_datetime,
                task_completed_date=task_completed_date,
                task_completed_time=task_completed_time,
                due_date=due_date,
                due_time=due_time
            )
            
            import os
            from django.utils import timezone
            
            images = request.FILES.getlist('image')
            for index, image in enumerate(images):
                filename = f"delivery_{delivery_task.id}_create_{int(timezone.now().timestamp())}_{index}{os.path.splitext(image.name)[1]}"
                image.name = filename
                DeliveryTaskImage.objects.create(
                    delivery_task=delivery_task,
                    image=image
                )
            
            messages.success(request, 'Delivery task created successfully!')
            return redirect('dashboard-delivery-task-list')
            
        except Exception as e:
            messages.error(request, f'Error creating task: {str(e)}')
            return self.get(request)
            
class DeliveryTaskDetailView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request, task_id):
        try:
            task = Task.objects.select_related(
                'employee',
                'delivery_details'
            ).prefetch_related(
                'delivery_details__images'
            ).get(
                id=task_id,
                task_type='delivery',
                employee__is_superuser=False,
                employee__is_staff=False,
                employee__is_deleted=False
            )
            
            context = {
                'task': task,
                'delivery': task.delivery_details
            }
            
            return render(request, 'delivery_task_detail.html', context)
            
        except Task.DoesNotExist:
            messages.error(request, 'Delivery Task not found')
            return redirect('dashboard-delivery-task-list')


            
class EditDeliveryTaskView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request, task_id):
        try:
            task = Task.objects.select_related(
                'employee',
                'delivery_details'
            ).get(
                id=task_id,
                task_type='delivery',
                employee__is_superuser=False,
                employee__is_staff=False,
                employee__is_deleted=False
            )
            
            context = {
                'task': task,
                'delivery': task.delivery_details,
                'employees': Employee.objects.filter(is_active=True, is_superuser=False, is_staff=False, is_deleted=False, employee_type='deliver'),
                'status_choices': Task.TASK_STATUS_CHOICES,
                'priority_choices': Task.PRIORITY_CHOICES,
            }
            return render(request, 'edit_delivery_task.html', context)
        except Task.DoesNotExist:
            messages.error(request, 'Delivery Task not found')
            return redirect('dashboard-delivery-task-list')
    
    def post(self, request, task_id):
        try:
            task = Task.objects.get(id=task_id)
            delivery = task.delivery_details
            
            # 1. Update Base Task
            task.employee_id = request.POST.get('employee')
            task.priority = request.POST.get('priority', 'medium')
            task.save()
            
            # 2. Update Delivery Specific Fields
            task_assign_datetime = request.POST.get('task_assign_datetime') or None
            
            due_date = request.POST.get('due_date') or None
            due_time = request.POST.get('due_time') or None

            # Update DeliveryTask
            delivery.DeliveryId = request.POST.get('DeliveryId')
            delivery.customer_name = request.POST.get('customer_name')
            delivery.customer_phone = request.POST.get('customer_phone')
            delivery.delivery_location = request.POST.get('delivery_location')
            delivery.delivery_notes = request.POST.get('delivery_notes')
            delivery.task_assign_datetime = task_assign_datetime
            delivery.due_date = due_date
            delivery.due_time = due_time
            delivery.save()
            
            import os
            from django.utils import timezone
            if request.FILES.getlist('image'):
                for index, image in enumerate(request.FILES.getlist('image')):
                    filename = f"delivery_{delivery.id}_edit_{int(timezone.now().timestamp())}_{index}{os.path.splitext(image.name)[1]}"
                    image.name = filename
                    DeliveryTaskImage.objects.create(
                        delivery_task=delivery,
                        image=image
                    )
            
            messages.success(request, 'Delivery task updated successfully!')
            return redirect('dashboard-delivery-task-list')
            
        except Exception as e:
            messages.error(request, f'Error updating task: {str(e)}')
            return self.get(request, task_id)


class MechanicTaskDetailView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request, task_id):
        try:
            task = Task.objects.select_related(
                'employee',
                'mechanic_details'
            ).get(
                id=task_id,
                task_type='mechanic',
                employee__is_superuser=False,
                employee__is_staff=False,
                employee__is_deleted=False
            )
            
            context = {
                'task': task,
                'mechanic': task.mechanic_details
            }
            
            return render(request, 'mechanic_task_detail.html', context)
            
        except Task.DoesNotExist:
            messages.error(request, 'Mechanic Task not found')
            return redirect('dashboard-mechanic-task-list')


class CreateMechanicTaskView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request):
        context = {
            'employees': Employee.objects.filter(is_active=True, is_superuser=False, is_staff=False, is_deleted=False, employee_type='mechanic'),
            'priority_choices': Task.PRIORITY_CHOICES,
            'part_number_choices': PartNumber.objects.all(),
        }
        return render(request, 'create_mechanic_task.html', context)
    
    def post(self, request):
        try:
            is_maintenance = request.POST.get('is_maintenance') == 'on'
            employee_id = request.POST.get('employee')
            
            if is_maintenance:
                employee_id = None
                
            priority = request.POST.get('priority', 'medium')
            
            # 1. Create Base Task
            task = Task.objects.create(
                employee_id=employee_id,
                task_type='mechanic',
                status='not_started',
                priority=priority,
                icon_type='mechanic',
                is_maintenance=is_maintenance
            )
            
            # 2. Create Mechanic Task Details
            started_at_str = request.POST.get('started_at')
            started_at = parse_datetime(started_at_str) if started_at_str else None
            
            mechanic_details = Mechanic.objects.create(
                task=task,
                heading=request.POST.get('heading'),
                Site_number=request.POST.get('Site_number'),
                bay_number=request.POST.get('bay_number'),
                Machine_type=request.POST.get('Machine_type'),
                Machine_serial_number=request.POST.get('Machine_serial_number'),
                job_description=request.POST.get('job_description'),
                spare_part_details=request.POST.get('spare_part_details'),
                started_at=started_at
            )
            
            # 3. Create Part Items if provided
            part_items_json = request.POST.get('part_items', '[]')
            try:
                part_items = json.loads(part_items_json) if part_items_json else []
                for item in part_items:
                    part_number_id = item.get('part_number_id') or item.get('part_number')
                    quantity = item.get('quantity', 1)
                    if part_number_id:
                        try:
                            part_num = PartNumber.objects.get(id=part_number_id)
                            MechanicPartItem.objects.create(
                                mechanic=mechanic_details,
                                part_number=part_num,
                                quantity=quantity
                            )
                        except PartNumber.DoesNotExist:
                            pass
            except (json.JSONDecodeError, TypeError):
                pass
            
            messages.success(request, 'Mechanic task created successfully!')
            return redirect('dashboard-mechanic-task-list')
            
        except Exception as e:
            messages.error(request, f'Error creating task: {str(e)}')
            return self.get(request)


class EditMechanicTaskView(LoginRequiredMixin, View):
    login_url = '/admin-login/'
    
    def get(self, request, task_id):
        try:
            task = Task.objects.select_related(
                'employee',
                'mechanic_details'
            ).get(
                Q(employee__isnull=True) | Q(employee__is_superuser=False, employee__is_staff=False, employee__is_deleted=False),
                id=task_id,
                task_type='mechanic',
            )
            
            context = {
                'task': task,
                'mechanic': task.mechanic_details,
                'employees': Employee.objects.filter(is_active=True, is_superuser=False, is_staff=False, is_deleted=False, employee_type='mechanic'),
                'status_choices': Task.TASK_STATUS_CHOICES,
                'priority_choices': Task.PRIORITY_CHOICES,
    
                'part_number_choices': PartNumber.objects.all(),
            }
            return render(request, 'edit_mechanic_task.html', context)
        except Task.DoesNotExist:
            messages.error(request, 'Mechanic Task not found')
            return redirect('dashboard-mechanic-task-list')
    
    def post(self, request, task_id):
        try:
            task = Task.objects.get(id=task_id)
            mechanic = task.mechanic_details
            
            # 1. Update Base Task
            is_maintenance = request.POST.get('is_maintenance') == 'on'
            if is_maintenance:
                task.employee = None
            else:
                task.employee_id = request.POST.get('employee')
                
            task.is_maintenance = is_maintenance
            task.priority = request.POST.get('priority', 'medium')
            task.status = request.POST.get('status', 'not_started')
            task.save()
            
            # 2. Update Mechanic Details
            mechanic.heading = request.POST.get('heading')
            mechanic.Site_number = request.POST.get('Site_number')
            mechanic.bay_number = request.POST.get('bay_number')
            mechanic.Machine_type = request.POST.get('Machine_type')
            mechanic.Machine_serial_number = request.POST.get('Machine_serial_number')
            mechanic.job_description = request.POST.get('job_description')
            mechanic.spare_part_details = request.POST.get('spare_part_details')
            
            # Update Timestamps
            started_at_str = request.POST.get('started_at')
            completed_at_str = request.POST.get('completed_at')
            
            if started_at_str:
                mechanic.started_at = parse_datetime(started_at_str)
            
            if completed_at_str:
                mechanic.completed_at = parse_datetime(completed_at_str)
            
            # Auto-update logic (as fallback)
            if task.status == 'in_progress' and not mechanic.started_at:
                mechanic.started_at = timezone.now()
            
            if task.status == 'completed' and not mechanic.completed_at:
                 mechanic.completed_at = timezone.now()
            elif task.status != 'completed' and not completed_at_str:
                 # Only nullify if no manual input was provided and status is not completed
                 mechanic.completed_at = None

            mechanic.completion_remarks = request.POST.get('completion_remarks')
            
            if request.FILES.get('completion_image'):
                 mechanic.completion_image = request.FILES.get('completion_image')
            
            mechanic.save()
            
            # 3. Update Part Items - delete old and create new
            part_items_json = request.POST.get('part_items', '[]')
            try:
                part_items = json.loads(part_items_json) if part_items_json else []
                # Delete existing part items
                mechanic.part_items.all().delete()
                # Create new part items
                for item in part_items:
                    part_number_id = item.get('part_number_id') or item.get('part_number')
                    quantity = item.get('quantity', 1)
                    if part_number_id:
                        try:
                            part_num = PartNumber.objects.get(id=part_number_id)
                            MechanicPartItem.objects.create(
                                mechanic=mechanic,
                                part_number=part_num,
                                quantity=quantity
                            )
                        except PartNumber.DoesNotExist:
                            pass
            except (json.JSONDecodeError, TypeError):
                pass
            
            messages.success(request, 'Mechanic task updated successfully!')
            return redirect('dashboard-mechanic-task-list')
            
        except Exception as e:
            messages.error(request, f'Error updating task: {str(e)}')
            return self.get(request, task_id)

class DeliveryNoteListView(LoginRequiredMixin, ListView):
    model = DeliveryNote
    template_name = 'dashboard/delivery_note_list.html'
    context_object_name = 'notes'
    paginate_by = 20

    def get_queryset(self):
        queryset = DeliveryNote.objects.select_related(
            'delivery_task', 
            'delivery_task__task',
            'delivery_task__task__employee',
            'delivery_task__task__employee__company'
        ).all().order_by('-created_at')
        
        # Filter by Employee (through delivery_task -> task -> employee)
        employee_id = self.request.GET.get('employee')
        if employee_id:
            queryset = queryset.filter(delivery_task__task__employee__id=employee_id)

        # Filter by Company (through delivery_task -> task -> employee -> company)
        company_id = self.request.GET.get('company')
        if company_id:
            queryset = queryset.filter(delivery_task__task__employee__company__id=company_id)

        # Search by employee name, customer name, or note content
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(delivery_task__task__employee__employee_name__icontains=search_query) |
                Q(delivery_task__task__employee__employeeId__icontains=search_query) |
                Q(delivery_task__customer_name__icontains=search_query) |
                Q(note__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['employees'] = Employee.objects.filter(
            is_active=True, 
            employee_type='deliver'
        ).order_by('employee_name')
        context['companies'] = Company.objects.filter(is_active=True).order_by('company_name')
        return context


class ClearAllNotificationsView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        try:
            # Clear all notifications for the current user
            Notification.objects.filter(recipient=request.user).delete()
            
            return JsonResponse({
                'success': True,
                'message': 'All notifications cleared successfully'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

